#!/usr/bin/env python3
"""
================================================================================
Benchmark Analysis Tool - Version 2.9
================================================================================
Script to process multiple Excel files and generate comparison with version tracking.

Version History:
 v1.0: Initial release with CSV file inputs
 v2.0: Major redesign - switched to Excel file input, optional folders, 3 ratio columns
 v2.1: Multiple Excel files support, API version detection, cross-version comparisons
 v2.2: Reorganized column order - backend columns first, then ratio columns
 v2.3: Simplified column names when only one input file (no version suffix)
 v2.4: Backup original backend pages into output workbook as separate sheets
 v2.5: Separate comparison pages per version + dedicated cross-version page
 v2.6: Extract summary columns from existing comparison page and append to new pages
 v2.7: Report issues AND filter benchmarks to only include those existing in all backends
 v2.8: Detect Skia version (m[0-9]+) and intelligently determine baseline/compare versions
 v2.9: Cache backend name extraction, simplify baseline handling, improve performance

Features:
 Accepts 0-2 folder paths for trace analysis (optional)
 Accepts MULTIPLE Excel files with multiple sheets (one per backend)
 Detects API version (api[0-9]+) AND Skia version (m[0-9]+) from filename
 Intelligently determines baseline and compare versions
 Generates comparison pages with baseline as reference
 Auto-increments output filename version number with baseline info
 Caches backend name extraction for better performance

Usage: python script.py [<folder_path1>] [<folder_path2>] <excel_file1.xlsx> [<excel_file2.xlsx> ...]

Author: Benchmark Analysis Tool
Version: 2.9
Date: 2026-06-15
================================================================================
"""

import sys
import os
import json
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.utils import get_column_letter
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# Version constant
MAJOR_VERSION = 2
MINOR_VERSION = 9
VERSION = f"{MAJOR_VERSION}.{MINOR_VERSION}"

# Global variable for API version prefixes
API_PREFIXES = ['api', 'r']
SKIA_PREFIXES = ['m']

# Build regex pattern dynamically from API_PREFIXES
API_PATTERN = '|'.join(re.escape(prefix) for prefix in API_PREFIXES)
SKIA_PATTERN = '|'.join(re.escape(prefix) for prefix in SKIA_PREFIXES)

API_REGEX = re.compile(f'({API_PATTERN})(\\d+)', re.IGNORECASE)
SKIA_REGEX = re.compile(f'({SKIA_PATTERN})(\\d+)', re.IGNORECASE)

def extract_versions_from_filename(filename):
    """Extract API version (api[0-9]+ or r[0-9]+) and Skia version (m[0-9]+) from filename."""
    api_version = None
    skia_version = None
    
    # Extract API version - uses global API_REGEX
    api_match = API_REGEX.search(filename)
    if api_match:
        prefix = api_match.group(1).lower()
        number = api_match.group(2)
        api_version = f"{prefix}{number}"
    
    # Extract Skia version - uses global SKIA_REGEX
    skia_match = SKIA_REGEX.search(filename)
    if skia_match:
        prefix = skia_match.group(1).lower()
        number = skia_match.group(2)
        skia_version = f"{prefix}{number}"
    
    return api_version, skia_version

def determine_baseline_and_compare(version_groups):
    """
    Determine which version is baseline and which is compare.
    Rules:
    - If there's only one version, use it as baseline and extract its components
    - If input files have same Skia version, use Skia version as baseline, API version as compare
    - If input files have same API version, use API version as baseline, Skia version as compare
    - If neither, use the first version as baseline
    """
    versions = list(version_groups.keys())
    
    # Handle single version case
    if len(versions) <= 1:
        full_version = versions[0]
        # Extract API and Skia from the single version
        api_match = API_REGEX.search(full_version)
        skia_match = SKIA_REGEX.search(full_version)
        
        if api_match and skia_match:
            # Both API and Skia present - use Skia as baseline, API as compare
            api_prefix = api_match.group(1).lower()
            api_number = api_match.group(2)
            skia_prefix = skia_match.group(1).lower()
            skia_number = skia_match.group(2)
            baseline_version = f"{skia_prefix}{skia_number}"
            compare_versions = [f"{api_prefix}{api_number}"]
            comparison_type = "single_with_both"
            print(f"\n📌 Single version detected: {full_version}")
            print(f"   Using Skia version as baseline: {baseline_version}")
            print(f"   API version as compare: {compare_versions}")
        elif api_match:
            # Only API present
            prefix = api_match.group(1).lower()
            number = api_match.group(2)
            baseline_version = f"{prefix}{number}"
            compare_versions = [baseline_version]
            comparison_type = "single_api_only"
            print(f"\n📌 Single version detected: {full_version}")
            print(f"   Using API version as baseline: {baseline_version}")
        elif skia_match:
            # Only Skia present
            prefix = skia_match.group(1).lower()
            number = skia_match.group(2)
            baseline_version = f"{prefix}{number}"
            compare_versions = [baseline_version]
            comparison_type = "single_skia_only"
            print(f"\n📌 Single version detected: {full_version}")
            print(f"   Using Skia version as baseline: {baseline_version}")
        else:
            # No version detected
            baseline_version = full_version
            compare_versions = [baseline_version]
            comparison_type = "single_no_version"
            print(f"\n📌 Single version detected: {full_version}")
            print(f"   Using full version as baseline: {baseline_version}")
        
        return baseline_version, compare_versions, comparison_type
    
    # Extract components from version strings for multiple versions
    version_components = {}
    for version in versions:
        api_match = API_REGEX.search(version)
        skia_match = SKIA_REGEX.search(version)
        version_components[version] = {
            'api': f"{api_match.group(1).lower()}{api_match.group(2)}" if api_match else None,
            'skia': f"{skia_match.group(1).lower()}{skia_match.group(2)}" if skia_match else None,
            'full': version
        }
    
    # Check if all versions have same Skia version
    skia_versions = set()
    for v in version_components.values():
        if v['skia']:
            skia_versions.add(v['skia'])
    
    # Check if all versions have same API version
    api_versions = set()
    for v in version_components.values():
        if v['api']:
            api_versions.add(v['api'])
    
    # Case: All files have the SAME Skia version (regardless of API)
    if len(skia_versions) == 1 and len(versions) > 1:
        # Same Skia version across all files - use Skia as baseline
        baseline_skia = list(skia_versions)[0]
        baseline_version = baseline_skia
        
        # Compare versions are the API versions (extract from full version)
        compare_versions = []
        for version in versions:
            api_match = API_REGEX.search(version)
            if api_match:
                compare_versions.append(f"{api_match.group(1).lower()}{api_match.group(2)}")
            else:
                compare_versions.append(version)
        # Remove duplicates
        seen = set()
        compare_versions = [x for x in compare_versions if not (x in seen or seen.add(x))]
        
        # Determine if API versions are also the same
        if len(api_versions) == 1:
            comparison_type = "same_skia_same_api"
            print(f"\n📌 Detected: Same Skia version ({baseline_skia}) AND same API version across files")
        else:
            comparison_type = "same_skia_different_api"
            print(f"\n📌 Detected: Same Skia version ({baseline_skia}) across files")
        
        print(f"   Using Skia version as baseline, compare versions: {compare_versions}")
        return baseline_version, compare_versions, comparison_type
    
    # Case: All files have the SAME API version (different Skia)
    if len(api_versions) == 1 and len(versions) > 1:
        # Same API version across all files - use API as baseline
        baseline_api = list(api_versions)[0]
        baseline_version = baseline_api
        
        # Compare versions are the Skia versions (extract from full version)
        compare_versions = []
        for version in versions:
            skia_match = SKIA_REGEX.search(version)
            if skia_match:
                compare_versions.append(f"{skia_match.group(1).lower()}{skia_match.group(2)}")
            else:
                compare_versions.append(version)
        # Remove duplicates
        seen = set()
        compare_versions = [x for x in compare_versions if not (x in seen or seen.add(x))]
        
        comparison_type = "same_api_different_skia"
        print(f"\n📌 Detected: Same API version ({baseline_api}) across files")
        print(f"   Using API version as baseline, compare versions: {compare_versions}")
        return baseline_version, compare_versions, comparison_type
    
    # Default: use first version as baseline
    baseline_version = versions[0]
    compare_versions = versions[1:]
    comparison_type = "mixed_versions"
    print(f"\n📌 No clear version pattern detected")
    print(f"   Using '{baseline_version}' as baseline")
    return baseline_version, compare_versions, comparison_type

def get_next_version_number(base_filename):
    """Get the next available version number for output filename."""
    pattern = re.compile(rf'{re.escape(base_filename)}_v(\d+)\.xlsx$')
    existing_versions = []
    
    # Check for existing files with version numbers
    for file in Path('.').glob(f'{base_filename}_v*.xlsx'):
        match = pattern.match(file.name)
        if match:
            existing_versions.append(int(match.group(1)))
    
    if existing_versions:
        next_version = max(existing_versions) + 1
    else:
        next_version = 1
    
    return next_version

def generate_output_filename(baseline_version):
    """Generate output filename with baseline version."""
    # Create base name: [baseline_version]_crossplatform_comparison
    base_name = f"{baseline_version}_crossplatform_comparison"
    
    # Get next version number
    version_num = get_next_version_number(base_name)
    
    # Create filename
    output_file = f"{base_name}_v{version_num}.xlsx"
    
    return output_file, version_num

def validate_arguments():
    """Validate command line arguments."""
    if len(sys.argv) < 2:
        print("Error: At least one argument required (Excel file)")
        print(f"Usage: {sys.argv[0]} <excel_file1.xlsx> [<excel_file2.xlsx> ...]")
        print("Note: At least one Excel file parameter.")
        sys.exit(1)
    
    # Parse arguments: first 0-2 could be folders, rest are Excel files
    excel_files = []
    
    # Check each argument
    for arg in sys.argv[1:]:
        if arg.lower().endswith('.xlsx') or arg.lower().endswith('.xls'):
            excel_files.append(arg)
        else:
            print(f"Error: No other parameters allowed. Extra parameter: '{arg}'")
            sys.exit(1)
    
    if len(excel_files) == 0:
        print("Error: At least one Excel file (.xlsx or .xls) must be provided")
        print(f"Usage: {sys.argv[0]} [<folder_path1>] [<folder_path2>] <excel_file1.xlsx> [<excel_file2.xlsx> ...]")
        sys.exit(1)
    
    # Check Excel files exist
    for excel_file in excel_files:
        if not os.path.isfile(excel_file):
            print(f"Error: Excel file '{excel_file}' does not exist")
            sys.exit(1)
    
    return excel_files

def detect_backend_type(folder_path):
    """Detect backend type from folder name."""
    folder_name = Path(folder_path).name.lower()
    if 'gr' in folder_name:
        return 'graphite'
    else:
        return 'ganesh'

def extract_summary_columns_from_comparison(excel_file):
    """Extract summary columns from existing comparison sheet if it exists."""
    try:
        # Load the workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Check if comparison sheet exists
        if 'comparison' not in wb.sheetnames:
            print(f"  No 'comparison' sheet found in {Path(excel_file).name}")
            return None
        
        # Read the comparison sheet
        df = pd.read_excel(excel_file, sheet_name='comparison')
        
        # Find columns that are summary columns (contain 'summary' in name or are not standard columns)
        standard_columns = ['ID', 'Bench']
        summary_columns = {}
        
        for col in df.columns:
            if col not in standard_columns:
                # Check if it looks like a summary column
                if 'summary' in col.lower() or 'draw' in col.lower() or 'trace' in col.lower():
                    summary_columns[col] = df[col].tolist()
                    print(f"  - Found summary column: '{col}'")
        
        if summary_columns:
            print(f"  Extracted {len(summary_columns)} summary columns from existing comparison page")
            return summary_columns
        else:
            print(f"  No summary columns found in existing comparison page")
            return None
            
    except Exception as e:
        print(f"  Error reading comparison sheet from {Path(excel_file).name}: {e}")
        return None

def extract_summary_columns_for_version(excel_file, version_tag):
    """Extract summary columns from comparison page for a specific version."""
    try:
        # Load the workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Check if comparison sheet exists
        if 'comparison' not in wb.sheetnames:
            return None
        
        # Read the comparison sheet
        df = pd.read_excel(excel_file, sheet_name='comparison')
        
        # Find columns that are summary columns
        standard_columns = ['ID', 'Bench']
        summary_columns = {}
        
        for col in df.columns:
            if col not in standard_columns:
                # Check if it looks like a summary column
                if 'summary' in col.lower() or 'draw' in col.lower() or 'trace' in col.lower():
                    summary_columns[col] = df[col].tolist()
                    print(f"    - Found summary column for version {version_tag}: '{col}'")
        
        if summary_columns:
            print(f"    Extracted {len(summary_columns)} summary columns from comparison page for version {version_tag}")
            return summary_columns
        else:
            return None
            
    except Exception as e:
        print(f"    Error reading comparison sheet from {Path(excel_file).name}: {e}")
        return None

def check_columns_coverage(version_groups, all_backends):
    """Check and report columns (backends) that don't exist in all versions."""
    print("\n" + "="*60)
    print("📊 Column/Backend Coverage Analysis")
    print("="*60)
    
    # For each backend, track which versions have it
    backend_coverage = defaultdict(list)
    
    for version_name, version_data in version_groups.items():
        # Get all backend names (without version suffix) for this version
        version_backends = set()
        for col_name in version_data['dataframes'].keys():
            # Get backend name from mapping
            backend_name = version_data['backend_mapping'].get(col_name, col_name)
            version_backends.add(backend_name)
        
        for backend in version_backends:
            backend_coverage[backend].append(version_name)
    
    # Find backends missing from some versions
    all_versions = set(version_groups.keys())
    missing_backends = {}
    
    for backend, present_versions in backend_coverage.items():
        present_set = set(present_versions)
        missing = all_versions - present_set
        if missing:
            missing_backends[backend] = {
                'present_in': sorted(present_versions),
                'missing_in': sorted(missing)
            }
    
    if missing_backends:
        print("\n⚠️  Backends missing from some versions:")
        for backend, info in sorted(missing_backends.items()):
            print(f"\n  Backend '{backend}':")
            print(f"    - Present in: {', '.join(info['present_in'])}")
            print(f"    - Missing in: {', '.join(info['missing_in'])}")
    else:
        print("\n✅ All backends are present in all versions!")
    
    # Also find backends that are only in a single version
    single_version_backends = {b: v for b, v in backend_coverage.items() if len(v) == 1}
    if single_version_backends:
        print(f"\n📌 Backends unique to a single version ({len(single_version_backends)}):")
        for backend, versions in sorted(single_version_backends.items()):
            print(f"    - '{backend}' (only in {versions[0]})")
    
    print("="*60)
    
    return missing_backends, single_version_backends

def check_duplicate_benchmarks(version_groups):
    """Check and report duplicate benchmark names within each sheet."""
    print("\n" + "="*60)
    print("🔍 Duplicate Benchmark Detection")
    print("="*60)
    
    duplicates_found = False
    duplicate_report = {}
    
    for version_name, version_data in version_groups.items():
        print(f"\n  Checking version: {version_name}")
        version_duplicates = {}
        
        for sheet_name, df in version_data['sheets'].items():
            # Check for duplicates in the bench column
            bench_counts = df['bench'].value_counts()
            duplicates = bench_counts[bench_counts > 1]
            
            if len(duplicates) > 0:
                duplicates_found = True
                version_duplicates[sheet_name] = duplicates.to_dict()
                print(f"    ⚠️  Sheet '{sheet_name}' has {len(duplicates)} duplicate benchmark(s):")
                for bench_name, count in duplicates.items():
                    print(f"        - '{bench_name}' appears {count} times")
            else:
                print(f"    ✓ Sheet '{sheet_name}' has no duplicate benchmarks")
        
        if version_duplicates:
            duplicate_report[version_name] = version_duplicates
    
    if not duplicates_found:
        print("\n✅ No duplicate benchmarks found in any sheet!")
    
    print("="*60)
    
    return duplicate_report

def check_missing_benchmarks(version_groups):
    """Check and report benchmark cases that don't exist in all pages."""
    print("\n" + "="*60)
    print("📋 Benchmark Coverage Analysis")
    print("="*60)
    
    # Get all unique benchmarks across all versions
    all_benchmarks = set()
    for version_data in version_groups.values():
        for df in version_data['dataframes'].values():
            all_benchmarks.update(df['bench'].tolist())
    
    # For each version, track which benchmarks are present
    version_benchmarks = {}
    for version_name, version_data in version_groups.items():
        benchmarks = set()
        for df in version_data['dataframes'].values():
            benchmarks.update(df['bench'].tolist())
        version_benchmarks[version_name] = benchmarks
    
    # Find benchmarks missing from each version
    missing_report = {}
    for version_name, benchmarks in version_benchmarks.items():
        missing = all_benchmarks - benchmarks
        if missing:
            missing_report[version_name] = missing
    
    if missing_report:
        print("\n⚠️  Missing Benchmarks Detected:")
        for version_name, missing in sorted(missing_report.items()):
            print(f"\n  Version '{version_name}' is missing {len(missing)} benchmark(s):")
            for bench in sorted(missing):
                # Find which versions have this benchmark
                present_in = []
                for v, b in version_benchmarks.items():
                    if bench in b:
                        present_in.append(v)
                if present_in:
                    print(f"    - {bench} (present in: {', '.join(present_in)})")
                else:
                    print(f"    - {bench} (present in: none)")
    else:
        print("\n✅ All benchmarks are present in all versions!")
    
    # Also check for benchmarks that are unique to a single version
    benchmark_versions = defaultdict(list)
    for version_name, benchmarks in version_benchmarks.items():
        for bench in benchmarks:
            benchmark_versions[bench].append(version_name)
    
    unique_benchmarks = {bench: versions for bench, versions in benchmark_versions.items() if len(versions) == 1}
    if unique_benchmarks:
        print(f"\n📌 Benchmarks unique to a single version ({len(unique_benchmarks)}):")
        for bench, versions in sorted(unique_benchmarks.items()):
            print(f"    - {bench} (only in {versions[0]})")
    
    print("="*60)
    
    return missing_report, unique_benchmarks, version_benchmarks

def read_excel_sheets(excel_file, version_tag):
    """Read all sheets from Excel file, excluding 'comparison' sheet."""
    try:
        # Dictionary to store dataframes with backend names
        dataframes = {}
        # Dictionary to store original sheet data for backup
        original_sheets = {}
        # Dictionary to store backend name mapping
        backend_mapping = {}
        
        # Load the workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Iterate through all sheets
        for sheet_name in wb.sheetnames:
            # Skip the 'comparison' sheet if it exists
            if sheet_name.lower() == 'comparison':
                print(f"  - Skipping existing 'comparison' sheet in {Path(excel_file).name}")
                continue
            
            # Read sheet into DataFrame
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Validate required columns
            if 'bench' not in df.columns or 'mean' not in df.columns:
                print(f"  Warning: Sheet '{sheet_name}' in {Path(excel_file).name} missing 'bench' or 'mean' column - skipping")
                continue
            
            # Store original sheet data for backup
            original_sheets[sheet_name] = df
            
            # Extract backend name by removing version suffixes (any part with digits)
            parts = sheet_name.split('_')
            backend_parts = []
            for part in parts:
                if not any(char.isdigit() for char in part):
                    backend_parts.append(part)
            backend_name = '_'.join(backend_parts)
            
            # Create column name with version tag
            column_name = f"{sheet_name}_{version_tag}" if version_tag else sheet_name
            dataframes[column_name] = df
            backend_mapping[column_name] = backend_name
            print(f"  - Loaded sheet '{sheet_name}' -> column '{column_name}' (backend: {backend_name}): {len(df)} rows")
        
        return dataframes, original_sheets, backend_mapping
        
    except Exception as e:
        print(f"Error reading Excel file '{excel_file}': {e}")
        return None, None, None

def import_existing_comparison_page(excel_file, version_tag, version_data):
    """Import existing comparison page from Excel file if it exists."""
    try:
        # Load the workbook with data_only=False to access formulas
        wb = load_workbook(excel_file, data_only=False)
        
        # Check if comparison sheet exists
        if 'comparison' not in wb.sheetnames:
            print(f"    No existing 'comparison' sheet found in {Path(excel_file).name}")
            return None
        
        # Get the comparison sheet
        sheet = wb['comparison']
        
        # Read the sheet with formulas
        # First, get all data as values
        data = []
        headers = []
        
        # Get headers from first row
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value is not None:
                headers.append(cell.value)
            else:
                headers.append(f"Unnamed: {col_idx}")
        
        # Read data rows
        for row in sheet.iter_rows(min_row=2, values_only=False):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    # Check if it's a formula
                    if cell.data_type == 'f':  # 'f' means formula
                        # Store as FORMULA: prefix with the formula string
                        row_data.append(f"FORMULA:{cell.value}")
                    else:
                        row_data.append(cell.value)
                else:
                    row_data.append(None)
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Check if there's already an average row
        has_average = False
        average_row_data = None
        
        if 'Bench' in df.columns:
            # Find rows where Bench contains 'AVERAGE' (case insensitive)
            avg_mask = df['Bench'].astype(str).str.upper() == 'AVERAGE'
            if avg_mask.any():
                has_average = True
                # Get the average row index
                avg_idx = avg_mask[avg_mask].index[0]
                average_row_data = df.iloc[avg_idx].to_dict()
                # Remove the average row from the dataframe
                df = df.drop(avg_idx).reset_index(drop=True)
                print(f"    Found existing average row - will preserve it")
        
        # Validate required columns
        if 'Bench' not in df.columns:
            print(f"    Warning: Existing comparison sheet missing 'Bench' column")
            return None
        
        # Clean up columns: Remove empty columns, unnamed columns, and columns that are entirely NaN
        columns_to_keep = []
        for col in df.columns:
            # Skip columns that are completely empty (all NaN or None)
            if df[col].isna().all() or df[col].isnull().all():
                print(f"    Skipping empty column: '{col}'")
                continue
            
            # Skip unnamed columns (typically from Excel where columns outside table area)
            if col.startswith('Unnamed:') or col == '' or col is None:
                print(f"    Skipping unnamed column: '{col}'")
                continue
            
            columns_to_keep.append(col)
        
        # Filter dataframe to only keep valid columns
        if columns_to_keep:
            df = df[columns_to_keep].copy()
            print(f"    Kept {len(columns_to_keep)} columns: {', '.join(columns_to_keep)}")
        else:
            print(f"    Warning: No valid columns found after cleaning")
            return None
        
        # Detect formula columns
        formula_columns = []
        if len(df) > 0:
            for col in df.columns:
                first_val = df[col].iloc[0]
                if isinstance(first_val, str) and first_val.startswith('FORMULA:'):
                    formula_columns.append(col)
        
        if formula_columns:
            print(f"    Detected formula columns: {', '.join(formula_columns)}")
        
        # Check if we need to filter benchmarks to those common across all backends
        dataframes = version_data['dataframes']
        backend_benchmarks = []
        for df_backend in dataframes.values():
            backend_benchmarks.append(set(df_backend['bench'].tolist()))
        
        if backend_benchmarks:
            common_benchmarks = set.intersection(*backend_benchmarks)
            # Filter the comparison dataframe to only include common benchmarks
            filtered_df = df[df['Bench'].isin(common_benchmarks)].copy()
            
            # Sort by Bench name for consistency
            filtered_df = filtered_df.sort_values('Bench').reset_index(drop=True)
            
            # Ensure ID column exists and is sequential
            if 'ID' in filtered_df.columns:
                # If ID exists, keep it but renumber
                filtered_df['ID'] = range(1, len(filtered_df) + 1)
            else:
                # Insert ID as first column
                filtered_df.insert(0, 'ID', range(1, len(filtered_df) + 1))
            
            print(f"    ✓ Imported existing comparison page: {len(filtered_df)} benchmarks (filtered to common backends)")
            print(f"    Columns preserved: {', '.join(filtered_df.columns.tolist())}")
            
            # Detect and report config columns
            config_columns = [col for col in filtered_df.columns if 'ratio' in col.lower() or 'diff' in col.lower() or 'vs' in col.lower()]
            if config_columns:
                print(f"    Detected config columns: {', '.join(config_columns)}")
            
            # If we have average row data, store it to be added back later
            if has_average and average_row_data is not None:
                # We'll add the average row back when writing
                filtered_df.attrs['has_average'] = True
                filtered_df.attrs['average_row'] = average_row_data
                print(f"    ✓ Will restore average row when writing")
            
            return filtered_df
        else:
            # If no filtering needed, just return the cleaned dataframe
            # Ensure ID column exists and is sequential
            if 'ID' in df.columns:
                df['ID'] = range(1, len(df) + 1)
            else:
                df.insert(0, 'ID', range(1, len(df) + 1))
            
            print(f"    ✓ Imported existing comparison page: {len(df)} benchmarks")
            print(f"    Columns preserved: {', '.join(df.columns.tolist())}")
            
            # Detect and report config columns
            config_columns = [col for col in df.columns if 'ratio' in col.lower() or 'diff' in col.lower() or 'vs' in col.lower()]
            if config_columns:
                print(f"    Detected config columns: {', '.join(config_columns)}")
            
            # If we have average row data, store it to be added back later
            if has_average and average_row_data is not None:
                # We'll add the average row back when writing
                df.attrs['has_average'] = True
                df.attrs['average_row'] = average_row_data
                print(f"    ✓ Will restore average row when writing")
            
            return df
            
    except Exception as e:
        print(f"    Error importing comparison sheet from {Path(excel_file).name}: {e}")
        return None

def check_existing_comparison_page(excel_file):
    """Check if a comparison page exists in the Excel file."""
    try:
        wb = load_workbook(excel_file, data_only=True)
        exists = 'comparison' in wb.sheetnames
        wb.close()
        return exists
    except Exception as e:
        return False
        
def read_multiple_excel_files(excel_files):
    """Read all Excel files and combine their data."""
    all_dataframes = {}
    all_backends = set()
    version_info = {}
    all_original_sheets = {}
    version_groups = defaultdict(dict)  # Group data by version tag
    summary_columns_by_version = {}  # Store summary columns per version
    has_comparison_page = {}  # Track which versions have existing comparison pages
    
    print(f"\n📖 Reading Excel files...")
    
    # Process each Excel file
    for excel_file in excel_files:
        filename = Path(excel_file).stem
        api_version, skia_version = extract_versions_from_filename(filename)
        
        # Create version tag (prioritize showing both if available)
        if api_version and skia_version:
            version_tag = f"{api_version}_{skia_version}"
            print(f"\n📖 Reading Excel file: {Path(excel_file).name} (API: {api_version}, Skia: {skia_version})")
        elif api_version:
            version_tag = api_version
            print(f"\n📖 Reading Excel file: {Path(excel_file).name} (API: {api_version})")
        elif skia_version:
            version_tag = skia_version
            print(f"\n📖 Reading Excel file: {Path(excel_file).name} (Skia: {skia_version})")
        else:
            version_tag = "default"
            print(f"\n📖 Reading Excel file: {Path(excel_file).name} (no version detected)")
        
        # Check if comparison page exists
        has_comparison = check_existing_comparison_page(excel_file)
        has_comparison_page[version_tag] = has_comparison
        if has_comparison:
            print(f"  ✓ Found existing 'comparison' page in this file")
        else:
            print(f"  ℹ️ No existing 'comparison' page found")
        
        # Extract summary columns from this file's comparison page
        print(f"  Checking for existing summary columns...")
        summary_columns = extract_summary_columns_for_version(excel_file, version_tag)
        if summary_columns:
            summary_columns_by_version[version_tag] = summary_columns
        
        # Read the sheets
        dataframes, original_sheets, backend_mapping = read_excel_sheets(excel_file, version_tag)
        
        if dataframes and original_sheets and backend_mapping:
            all_dataframes.update(dataframes)
            
            # Store original sheets with version info
            for sheet_name, df in original_sheets.items():
                backup_sheet_name = f"{sheet_name}_{version_tag}"
                all_original_sheets[backup_sheet_name] = df
            
            # Group data by version
            version_groups[version_tag]['dataframes'] = dataframes
            version_groups[version_tag]['file'] = excel_file
            version_groups[version_tag]['sheets'] = original_sheets
            version_groups[version_tag]['backend_mapping'] = backend_mapping
            version_groups[version_tag]['api'] = api_version
            version_groups[version_tag]['skia'] = skia_version
            version_groups[version_tag]['has_comparison_page'] = has_comparison
            
            # Store summary columns for this version
            if summary_columns:
                version_groups[version_tag]['summary_columns'] = summary_columns
            
            version_info[version_tag] = {
                'file': excel_file,
                'columns': list(dataframes.keys()),
                'api': api_version,
                'skia': skia_version,
                'backend_mapping': backend_mapping,
                'has_summary': summary_columns is not None,
                'has_comparison_page': has_comparison
            }
            
            # Collect unique backend names from the mapping
            for backend_name in backend_mapping.values():
                all_backends.add(backend_name)
    
    if len(all_dataframes) == 0:
        print("Error: No valid data loaded from any Excel file")
        sys.exit(1)
    
    # Determine baseline and compare versions
    baseline_version, compare_versions, comparison_type = determine_baseline_and_compare(version_groups)
    
    # Check for missing/extra backends across versions
    missing_backends, single_version_backends = check_columns_coverage(version_groups, all_backends)
    
    # Check for duplicate benchmarks
    duplicate_report = check_duplicate_benchmarks(version_groups)
    
    # Check for missing benchmarks
    missing_report, unique_benchmarks, version_benchmarks = check_missing_benchmarks(version_groups)
    
    # Determine which summary columns to use (prefer baseline version's summary if available)
    primary_summary_columns = None
    if baseline_version in summary_columns_by_version:
        primary_summary_columns = summary_columns_by_version[baseline_version]
        print(f"\n📋 Using summary columns from baseline version '{baseline_version}'")
    elif summary_columns_by_version:
        # Use the first available summary columns
        first_version = list(summary_columns_by_version.keys())[0]
        primary_summary_columns = summary_columns_by_version[first_version]
        print(f"\n📋 Using summary columns from version '{first_version}' (baseline has no summary)")
    
    print(f"\n📊 Summary: Loaded {len(all_dataframes)} backend columns from {len(excel_files)} files")
    print(f"   Backends found: {', '.join(sorted(all_backends))}")
    print(f"   Versions found: {', '.join(version_groups.keys())}")
    print(f"   Baseline version: {baseline_version}")
    print(f"   Compare versions: {', '.join(compare_versions) if compare_versions else 'None'}")
    print(f"   Original sheets to backup: {len(all_original_sheets)}")
    
    return (all_dataframes, all_backends, version_info, version_groups, 
            all_original_sheets, primary_summary_columns, missing_report, 
            unique_benchmarks, duplicate_report, missing_backends, 
            single_version_backends, version_benchmarks, baseline_version, 
            compare_versions, comparison_type, summary_columns_by_version, has_comparison_page)

def parse_summary_column(summary_data):
    """
    Parse summary column string format: 
    sub{idx}({sub_count}):[{draw_count}({mismatch_count})|{draw_type:type_count,...}]
    
    Multiple sub-entries can be separated by ',' or '\n'.
    Returns a list of draw type entries with their counts and weights.
    """
    draw_types = {}
    
    for entry in summary_data:
        if not isinstance(entry, str) or not entry.startswith('sub'):
            continue
        
        try:
            # Split multiple sub-entries by comma or newline
            # But carefully avoid splitting draw type pairs
            entry_cleaned = entry.replace('\n', ',')
            
            # Use a more sophisticated approach to split sub-entries
            # Find all 'sub' occurrences and split by them
            sub_entries = []
            current_pos = 0
            
            # Find all positions where 'sub' appears
            sub_positions = []
            for match in re.finditer(r'sub\d+\(', entry_cleaned):
                sub_positions.append(match.start())
            
            # If no sub positions found, return
            if not sub_positions:
                continue
            
            # Extract each sub-entry
            for i, pos in enumerate(sub_positions):
                if i < len(sub_positions) - 1:
                    # Extract from this sub to the next sub
                    next_pos = sub_positions[i + 1]
                    sub_entry = entry_cleaned[pos:next_pos].strip()
                else:
                    # Last sub-entry goes to the end
                    sub_entry = entry_cleaned[pos:].strip()
                
                if sub_entry.startswith('sub'):
                    sub_entries.append(sub_entry)
            
            if not sub_entries:
                continue
            
            # First pass: calculate total sub_count
            total_sub_count = 0
            for sub_entry in sub_entries:
                sub_match = re.search(r'sub\d+\((\d+)\)', sub_entry)
                if sub_match:
                    total_sub_count += int(sub_match.group(1))
            
            if total_sub_count == 0:
                continue
            
            # Second pass: parse each sub-entry
            for sub_entry in sub_entries:
                # Extract sub_count
                sub_match = re.search(r'sub\d+\((\d+)\)', sub_entry)
                if not sub_match:
                    continue
                sub_count = int(sub_match.group(1))
                
                # Extract draw data part after colon
                draw_part = sub_entry.split(':', 1)[1] if ':' in sub_entry else ''
                
                # Extract draw_count and optional mismatch_count
                draw_count_match = re.search(r'\[(\d+)\((\d+)\)\|', draw_part)
                if draw_count_match:
                    draw_count = int(draw_count_match.group(1))
                    mismatch_count = int(draw_count_match.group(2))
                else:
                    # No mismatch count
                    draw_count_match = re.search(r'\[(\d+)\|', draw_part)
                    if not draw_count_match:
                        continue
                    draw_count = int(draw_count_match.group(1))
                    mismatch_count = 0
                
                # Extract draw_type:count pairs
                draw_data = draw_part.split('|', 1)[1] if '|' in draw_part else ''
                
                # Remove the trailing ']' if present
                if draw_data.endswith(']'):
                    draw_data = draw_data[:-1]
                
                if not draw_data:
                    continue
                
                # Parse draw_type:count pairs (comma separated)
                pairs = draw_data.split(',')
                for pair in pairs:
                    pair = pair.strip()
                    if ':' in pair:
                        draw_type, count_str = pair.split(':', 1)
                        draw_type = draw_type.strip()
                        count_str = count_str.strip()
                        try:
                            count = int(count_str)
                            
                            # Calculate weight: (type_count/draw_count) * (sub_count/total_sub_count)
                            if draw_type not in draw_types:
                                draw_types[draw_type] = []
                            draw_types[draw_type].append({
                                'sub_count': sub_count,
                                'draw_count': draw_count,
                                'type_count': count,
                                'entry': entry
                            })
                        except ValueError:
                            # Skip if count is not a valid integer
                            continue
        except Exception as e:
            # Skip malformed entries
            continue
    
    return draw_types

def calculate_draw_type_weights_for_benchmark(summary_entry, total_sub_count):
    """
    Calculate weighted counts for each draw type for a single benchmark entry.
    Multiple sub-entries can be separated by ',' or '\n'.
    Returns a dictionary mapping draw_type to weighted count.
    """
    if not isinstance(summary_entry, str) or not summary_entry.startswith('sub'):
        return {}
    
    try:
        # Split multiple sub-entries by comma or newline
        # But carefully avoid splitting draw type pairs
        entry_cleaned = summary_entry.replace('\n', '')
        
        # Use a more sophisticated approach to split sub-entries
        # Find all 'sub' occurrences and split by them
        sub_entries = []
        current_pos = 0
        
        # Find all positions where 'sub' appears
        sub_positions = []
        for match in re.finditer(r'sub\d+\(', entry_cleaned):
            sub_positions.append(match.start())
        
        # If no sub positions found, return
        if not sub_positions:
            return {}
        
        # Extract each sub-entry
        for i, pos in enumerate(sub_positions):
            if i < len(sub_positions) - 1:
                # Extract from this sub to the next sub
                next_pos = sub_positions[i + 1]
                sub_entry = entry_cleaned[pos:next_pos-1].strip()
            else:
                # Last sub-entry goes to the end
                sub_entry = entry_cleaned[pos:].strip()
            
            if sub_entry.startswith('sub'):
                sub_entries.append(sub_entry)
        
        if not sub_entries:
            return {}
        
        weighted_counts = {}
        total_sub_count_actual = 0
        
        # First pass: calculate total sub_count
        for sub_entry in sub_entries:
            sub_match = re.search(r'sub\d+\((\d+)\)', sub_entry)
            if sub_match:
                total_sub_count_actual += int(sub_match.group(1))
        
        if total_sub_count_actual == 0:
            return {}
        
        # Second pass: calculate weights for each draw type
        for sub_entry in sub_entries:
            # Extract sub_count
            sub_match = re.search(r'sub\d+\((\d+)\)', sub_entry)
            if not sub_match:
                continue
            sub_count = int(sub_match.group(1))
            
            # Extract draw data part after colon
            draw_part = sub_entry.split(':', 1)[1] if ':' in sub_entry else ''
            
            # Extract draw_count and optional mismatch_count
            draw_count_match = re.search(r'\[(\d+)\((\d+)\)\|', draw_part)
            if draw_count_match:
                draw_count = int(draw_count_match.group(1))
                mismatch_count = int(draw_count_match.group(2))
            else:
                # No mismatch count
                draw_count_match = re.search(r'\[(\d+)\|', draw_part)
                if not draw_count_match:
                    continue
                draw_count = int(draw_count_match.group(1))
                mismatch_count = 0
            
            # Extract draw_type:count pairs
            draw_data = draw_part.split('|', 1)[1] if '|' in draw_part else ''
            
            # Remove the trailing ']' if present
            if draw_data.endswith(']'):
                draw_data = draw_data[:-1]
            
            if not draw_data:
                continue
            
            # Parse draw_type:count pairs (comma separated)
            pairs = draw_data.split(',')
            for pair in pairs:
                pair = pair.strip()
                if ':' in pair:
                    draw_type, count_str = pair.split(':', 1)
                    draw_type = draw_type.strip()
                    count_str = count_str.strip()
                    try:
                        count = int(count_str)
                        # Weight: (type_count/draw_count) * (sub_count/total_sub_count)
                        if draw_count > 0 and total_sub_count_actual > 0:
                            weight = (count / draw_count) * (sub_count / total_sub_count_actual)
                            if draw_type in weighted_counts:
                                weighted_counts[draw_type] += weight
                            else:
                                weighted_counts[draw_type] = weight
                    except ValueError:
                        # Skip if count is not a valid integer
                        continue
        
        if weighted_counts:
            total_weight = sum(weighted_counts.values())
            tolerance = 1e-9
            if abs(total_weight - 1.0) > tolerance:
                print(f"    ⚠️ WARNING: Weighted counts sum to {total_weight:.10f}, expected 1.0")
                # Normalize to ensure sum equals 1
        
        return weighted_counts
        
    except Exception as e:
        return {}

def get_draw_type_data_from_summary(df, summary_columns):
    """
    Extract draw type data from summary columns.
    Multiple sub-entries can be separated by ',' or '\n'.
    Returns a dictionary mapping:
    - summary_column_name -> {
        'backend': backend_name,
        'data': {benchmark: {draw_type: weighted_count, ...}}
      }
    """
    if not summary_columns:
        return {}
    
    draw_type_data = {}
    
    for col_name, col_values in summary_columns.items():
        # Extract backend name from column name
        # Column names like "grdawn_vk summary", "glesdmsaa summary", etc.
        backend_name = col_name.replace(' summary', '').replace(' Summary', '').strip()
        
        # Check if this column contains draw data
        has_draw_data = False
        for val in col_values[:10]:
            if isinstance(val, str) and 'sub' in val and ':' in val and '|' in val:
                has_draw_data = True
                break
        
        if not has_draw_data:
            continue
        
        # Get benchmarks from the df (excluding AVERAGE row)
        benches = df[df['Bench'] != 'AVERAGE']['Bench'].tolist() if 'Bench' in df.columns else []
        
        # Calculate total sub_count for each benchmark
        bench_data = {}
        
        for idx, bench in enumerate(benches):
            if idx < len(col_values):
                summary_entry = col_values[idx]
                if isinstance(summary_entry, str) and 'sub' in summary_entry:
                    # Calculate total sub_count for this benchmark
                    # Extract all sub_count values from the entry
                    # Handle both comma and newline separated sub-entries
                    entry_cleaned = summary_entry.replace('\n', ',')
                    
                    # Find all sub entries
                    sub_entries = []
                    sub_positions = []
                    for match in re.finditer(r'sub\d+\(', entry_cleaned):
                        sub_positions.append(match.start())
                    
                    if sub_positions:
                        for i, pos in enumerate(sub_positions):
                            if i < len(sub_positions) - 1:
                                next_pos = sub_positions[i + 1]
                                sub_entry = entry_cleaned[pos:next_pos].strip()
                            else:
                                sub_entry = entry_cleaned[pos:].strip()
                            if sub_entry.startswith('sub'):
                                sub_entries.append(sub_entry)
                    
                    if sub_entries:
                        total_sub_count = 0
                        for sub_entry in sub_entries:
                            sub_match = re.search(r'sub\d+\((\d+)\)', sub_entry)
                            if sub_match:
                                total_sub_count += int(sub_match.group(1))
                        
                        if total_sub_count > 0:
                            # Calculate weighted counts for each draw type
                            weighted_counts = calculate_draw_type_weights_for_benchmark(summary_entry, total_sub_count)
                            if weighted_counts:
                                bench_data[bench] = weighted_counts
        
        if bench_data:
            draw_type_data[col_name] = {
                'backend': backend_name,
                'data': bench_data
            }
            print(f"    ✓ Extracted draw type data from '{col_name}' (backend: {backend_name})")
    
    return draw_type_data

def calculate_draw_type_distribution(dist_df, df, summary_col_name, config_col_name):
    """
    Calculate and populate draw type weighted counts into the distribution table.
    
    This is the unique logic that:
    1. Parses summary strings to extract draw type weighted counts per benchmark
    2. Uses the config column (grdawn_vk vs glesdmsaa ratio/diff) to determine bin placement
    3. Handles both numeric values and formula expressions in the config column
    4. Aggregates weighted counts into the appropriate bin for each draw type
    
    Args:
        dist_df: The distribution DataFrame to fill (with Bin, Label, and draw type columns)
        df: The main comparison DataFrame
        summary_col_name: The summary column name containing draw type data
        config_col_name: The config column name (grdawn_vk vs glesdmsaa ratio/diff)
    
    Returns:
        Filled distribution DataFrame with weighted counts for each draw type per bin
    """
    if summary_col_name not in df.columns or config_col_name not in df.columns:
        return dist_df
    
    # Extract draw type data from the summary column
    summary_values = df[summary_col_name].tolist()
    draw_type_data = get_draw_type_data_from_summary(df, {summary_col_name: summary_values})
    
    if not draw_type_data or summary_col_name not in draw_type_data:
        return dist_df
    
    summary_info = draw_type_data[summary_col_name]
    bench_data = summary_info['data']
    
    # Get all unique draw types from this backend
    all_draw_types = set()
    for bench_info in bench_data.values():
        all_draw_types.update(bench_info.keys())
    
    if not all_draw_types:
        return dist_df
    
    sorted_draw_types = sorted(all_draw_types)
    
    # Ensure the distribution DataFrame has columns for all draw types
    for draw_type in sorted_draw_types:
        if draw_type not in dist_df.columns:
            dist_df[draw_type] = 0.0
    
    # Get bin thresholds from the distribution table
    bin_thresholds = []
    for val in dist_df['Bin']:
        if val != '' and not pd.isna(val):
            try:
                bin_thresholds.append(float(val))
            except (ValueError, TypeError):
                pass
    
    # Get the config values for each benchmark (excluding AVERAGE row)
    main_df = df[df['Bench'] != 'AVERAGE']
    config_values = main_df[config_col_name].tolist()
    benches = main_df['Bench'].tolist()
    
    # For each benchmark, get its weighted draw type counts and config value
    for bench_idx, bench in enumerate(benches):
        if bench_idx < len(config_values):
            config_val = config_values[bench_idx]
            
            # Skip invalid config values
            if config_val is None or pd.isna(config_val):
                continue
            
            # Calculate numeric value from formula if needed (private calculation)
            numeric_val = None
            
            # Check if it's a formula placeholder
            if isinstance(config_val, str) and config_val.startswith('FORMULA:'):
                formula_expr = config_val[8:]  # Remove 'FORMULA:' prefix
                                
                # Evaluate the formula using the actual data
                try:
                    # Parse formula expression like "grdawn_vk/glesdmsaa" or "grdawn_vk-glesdmsaa"
                    if '/' in formula_expr:
                        parts = formula_expr.split('/')
                        if len(parts) == 2:
                            num_ref = parts[0].strip()
                            den_ref = parts[1].strip()
                            
                            # Parse Excel cell reference (e.g., "D10" -> column D, row 10)
                            num_col_letter = ''.join([c for c in num_ref if c.isalpha()])
                            num_row = int(''.join([c for c in num_ref if c.isdigit()]))
                            
                            den_col_letter = ''.join([c for c in den_ref if c.isalpha()])
                            den_row = int(''.join([c for c in den_ref if c.isdigit()]))
                            
                            # Convert column letter to index (A=0, B=1, etc.)
                            num_col_idx = 0
                            for char in num_col_letter:
                                num_col_idx = num_col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
                            num_col_idx -= 1  # Convert to 0-based
                            
                            den_col_idx = 0
                            for char in den_col_letter:
                                den_col_idx = den_col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
                            den_col_idx -= 1  # Convert to 0-based
                            
                            # Get values from DataFrame
                            # Note: row_idx in DataFrame corresponds to Excel row (row_idx + 2 in Excel)
                            # But the formula references specific Excel rows, so we need to map them
                            # Excel row 1 = header, so data rows start at 2
                            # DataFrame index 0 = Excel row 2
                            # So DataFrame row index = Excel row - 2
                            df_row_idx = num_row - 2
                            den_df_row_idx = den_row - 2
                            
                            # Get column names from DataFrame
                            if num_col_idx < len(df.columns) and den_col_idx < len(df.columns):
                                num_col_name = df.columns[num_col_idx]
                                den_col_name = df.columns[den_col_idx]
                                
                                num_val = df.iloc[df_row_idx][num_col_name] if df_row_idx < len(df) else None
                                den_val = df.iloc[den_df_row_idx][den_col_name] if den_df_row_idx < len(df) else None
                                
                                if num_val is not None and den_val is not None and den_val != 0:
                                    try:
                                        numeric_val = float(num_val) / float(den_val)
                                    except (ValueError, TypeError, ZeroDivisionError):
                                        pass
                    elif '-' in formula_expr:
                        parts = formula_expr.split('-')
                        if len(parts) == 2:
                            num_ref = parts[0].strip()
                            den_ref = parts[1].strip()
                            
                            # Parse Excel cell reference
                            num_col_letter = ''.join([c for c in num_ref if c.isalpha()])
                            num_row = int(''.join([c for c in num_ref if c.isdigit()]))
                            
                            den_col_letter = ''.join([c for c in den_ref if c.isalpha()])
                            den_row = int(''.join([c for c in den_ref if c.isdigit()]))
                            
                            # Convert column letter to index
                            num_col_idx = 0
                            for char in num_col_letter:
                                num_col_idx = num_col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
                            num_col_idx -= 1
                            
                            den_col_idx = 0
                            for char in den_col_letter:
                                den_col_idx = den_col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
                            den_col_idx -= 1
                            
                            # Get values from DataFrame
                            df_row_idx = num_row - 2
                            den_df_row_idx = den_row - 2
                            
                            if num_col_idx < len(df.columns) and den_col_idx < len(df.columns):
                                num_col_name = df.columns[num_col_idx]
                                den_col_name = df.columns[den_col_idx]
                                
                                num_val = df.iloc[df_row_idx][num_col_name] if df_row_idx < len(df) else None
                                den_val = df.iloc[den_df_row_idx][den_col_name] if den_df_row_idx < len(df) else None
                                
                                if num_val is not None and den_val is not None:
                                    try:
                                        numeric_val = float(num_val) - float(den_val)
                                    except (ValueError, TypeError):
                                        pass
                except (ValueError, TypeError, IndexError, ZeroDivisionError):
                    pass
            
            # If not a formula or formula evaluation failed, try to convert directly
            if numeric_val is None:
                try:
                    numeric_val = float(config_val)
                except (ValueError, TypeError):
                    continue
            
            # Skip if we couldn't get a numeric value
            if numeric_val is None:
                continue
            
            # Determine which bin this config value falls into
            bin_idx = -1
            for i, threshold in enumerate(bin_thresholds):
                if numeric_val <= threshold:
                    bin_idx = i
                    break
            
            # If value is greater than all thresholds, it goes to the last bin
            if bin_idx == -1:
                bin_idx = len(bin_thresholds)  # Last bin (the one with empty Bin value)
            
            # Get weighted counts for this benchmark
            weighted_counts = bench_data.get(bench, {})
            
            # Add weighted counts to the appropriate bin and draw type cells
            for draw_type, weight in weighted_counts.items():
                if draw_type in dist_df.columns:
                    # Convert to float and add weight
                    current_val = dist_df.at[bin_idx, draw_type]
                    if isinstance(current_val, str):
                        current_val = 0.0
                    dist_df.at[bin_idx, draw_type] = float(current_val) + float(weight)
    
    # Round values to 3 decimal places for readability
    for col in sorted_draw_types:
        if col in dist_df.columns:
            dist_df[col] = dist_df[col].round(3)
    
    return dist_df

def add_glesdmsaa_distribution_tables(df, glesdmsaa_ratio_cols, glesdmsaa_diff_cols, version_name, summary_columns=None):
    """
    Add glesdmsaa distribution tables to the DataFrame as attributes.
    This function is used for both imported and generated data.
    
    Now includes draw call count distribution for each backend vs glesdmsaa config.
    """
    
    # Define standard bins (used for both glesdmsaa and draw distributions)
    ratio_bins = [
        (0.5, '[0, 0.5]'),
        (1.0, '[0.5, 1]'),
        (2.0, '[1, 2]'),
        (10.0, '[2, 10]'),
        (50.0, '[10, 50]'),
        (100.0, '[50, 100]'),
        (float('inf'), '>100')
    ]
    
    diff_bins = [
        (0.0, '<0'),
        (1.0, '[0, 1]'),
        (10.0, '[1, 10]'),
        (50.0, '[10, 50]'),
        (100.0, '[50, 100]'),
        (float('inf'), '>100')
    ]
    
    # ============================================================
    # Create glesdmsaa ratio distribution table
    # ============================================================
    if glesdmsaa_ratio_cols:        
        # Build ratio distribution table
        glesdmsaa_ratio_dist_data = {
            'Bin': [threshold for threshold, _ in ratio_bins if threshold != float('inf')] + [''],
            'Label': [label for _, label in ratio_bins]
        }
        
        # Add empty columns for each glesdmsaa ratio column
        for col in glesdmsaa_ratio_cols:
            glesdmsaa_ratio_dist_data[col] = [''] * len(ratio_bins)
        
        glesdmsaa_ratio_dist_df = pd.DataFrame(glesdmsaa_ratio_dist_data)
        
        # Store as attributes
        df.attrs['glesdmsaa_ratio_distribution'] = glesdmsaa_ratio_dist_df
        df.attrs['glesdmsaa_ratio_bins'] = [threshold for threshold, _ in ratio_bins if threshold != float('inf')]
        df.attrs['glesdmsaa_ratio_columns'] = glesdmsaa_ratio_cols
        
        print(f"    ✓ Created glesdmsaa ratio distribution table with {len(ratio_bins)} bins and {len(glesdmsaa_ratio_cols)} columns")
    
    # ============================================================
    # Create glesdmsaa diff distribution table
    # ============================================================
    if glesdmsaa_diff_cols:
        # Build diff distribution table
        glesdmsaa_diff_dist_data = {
            'Bin': [threshold for threshold, _ in diff_bins if threshold != float('inf')] + [''],
            'Label': [label for _, label in diff_bins]
        }
        
        # Add empty columns for each glesdmsaa diff column
        for col in glesdmsaa_diff_cols:
            glesdmsaa_diff_dist_data[col] = [''] * len(diff_bins)
        
        glesdmsaa_diff_dist_df = pd.DataFrame(glesdmsaa_diff_dist_data)
        
        # Store as attributes
        df.attrs['glesdmsaa_diff_distribution'] = glesdmsaa_diff_dist_df
        df.attrs['glesdmsaa_diff_bins'] = [threshold for threshold, _ in diff_bins if threshold != float('inf')]
        df.attrs['glesdmsaa_diff_columns'] = glesdmsaa_diff_cols
        
        print(f"    ✓ Created glesdmsaa diff distribution table with {len(diff_bins)} bins and {len(glesdmsaa_diff_cols)} columns")
    

    # ============================================================
    # Generate draw call distribution tables
    # ============================================================
    if summary_columns:
        print(f"    Generating draw call distribution tables from summary columns...")
        
        # Find config columns for grdawn_vk vs glesdmsaa
        ratio_config_col = None
        diff_config_col = None
        
        for col in df.columns:
            if "grdawn_vk vs glesdmsaa" in col:
                if "ratio" in col.lower():
                    ratio_config_col = col
                elif "diff" in col.lower():
                    diff_config_col = col
        
        # Process each summary column (each represents a different backend)
        for summary_col_name in summary_columns.keys():
            summary_type_name = summary_col_name.replace("summary of ", "")
            print(f"creating draw call distribution data for {summary_type_name}")

            # Generate ratio draw distribution using standard ratio_bins
            if ratio_config_col:
                # Build base distribution table structure
                bin_thresholds = [threshold for threshold, _ in ratio_bins if threshold != float('inf')]
                bin_labels = [label for _, label in ratio_bins]
                
                dist_data = {
                    'Bin': bin_thresholds + [''],
                    'Label': bin_labels
                }
                ratio_dist_df = pd.DataFrame(dist_data)
                                
                # Fill with weighted counts using the unique logic
                ratio_dist_df = calculate_draw_type_distribution(
                    ratio_dist_df, df, summary_col_name, ratio_config_col
                )
                
                # Check if any draw types were found
                draw_columns = [col for col in ratio_dist_df.columns if col not in ['Bin', 'Label']]
                if draw_columns:
                    # Store as attributes
                    attr_name = f"{summary_type_name}_ratio_draw_distribution"
                    df.attrs[attr_name] = ratio_dist_df
                    #df.attrs[f"{summary_type_name}_ratio_draw_types"] = sorted(draw_columns)
                    #df.attrs[f"{summary_type_name}_ratio_draw_config_column"] = ratio_config_col
                    #df.attrs[f"{summary_type_name}_ratio_draw_bins"] = ratio_bins
                    
                    print(f"    ✓ Created ratio draw distribution for summary column '{summary_type_name}' with {len(draw_columns)} draw types")
            
            # Generate diff draw distribution using standard diff_bins
            if diff_config_col:
                # Build base distribution table structure
                bin_thresholds = [threshold for threshold, _ in diff_bins if threshold != float('inf')]
                bin_labels = [label for _, label in diff_bins]
                
                dist_data = {
                    'Bin': bin_thresholds + [''],
                    'Label': bin_labels
                }
                diff_dist_df = pd.DataFrame(dist_data)
                
                # Fill with weighted counts using the unique logic
                diff_dist_df = calculate_draw_type_distribution(
                    diff_dist_df, df, summary_col_name, diff_config_col
                )
                
                # Check if any draw types were found
                draw_columns = [col for col in diff_dist_df.columns if col not in ['Bin', 'Label']]
                if draw_columns:
                    # Store as attributes
                    attr_name = f"{summary_type_name}_diff_draw_distribution"
                    df.attrs[attr_name] = diff_dist_df
                    #df.attrs[f"{summary_type_name}_diff_draw_types"] = sorted(draw_columns)
                    #df.attrs[f"{summary_type_name}_diff_draw_config_column"] = diff_config_col
                    #df.attrs[f"{summary_type_name}_diff_draw_bins"] = diff_bins
                    
                    print(f"    ✓ Created diff draw distribution for summary column '{summary_type_name}' with {len(draw_columns)} draw types")
        
        print(f"    ✓ Finished generating draw distribution tables")
    else:
        print(f"    ⚠️ No summary columns provided - cannot generate draw distribution tables")
    
    return df

def create_version_comparison_page(version_data, version_name,
                                    summary_columns, missing_benchmarks_for_version, 
                                    version_benchmarks, baseline_version, is_baseline=False):
    """Create comparison page for a specific version."""
    dataframes = version_data['dataframes']
    backend_mapping = version_data['backend_mapping']
    
    # Use version-specific summary columns if available, otherwise use the provided ones
    version_summary_columns = version_data.get('summary_columns', summary_columns)
    
    # Check if we have an existing comparison page to import
    has_existing = version_data.get('has_comparison_page', False)
    excel_file = version_data.get('file')
    
    # ============================================================
    # CASE 1: Import existing comparison page
    # ============================================================
    if has_existing and excel_file:
        print(f"    Found existing comparison page in {Path(excel_file).name} - importing it")
        imported_df = import_existing_comparison_page(excel_file, version_name, version_data)
        if imported_df is not None and not imported_df.empty:
            print(f"    ✓ Successfully imported existing comparison page with {len(imported_df)} benchmarks")
            print(f"    ✓ Preserved all existing columns including configs")
            
            # ============================================================
            # Identify glesdmsaa columns from imported data
            # ============================================================
            glesdmsaa_ratio_cols = []
            glesdmsaa_diff_cols = []
            
            # Search for columns that contain 'glesdmsaa' in their name
            for col in imported_df.columns:
                if 'glesdmsaa' in col.lower():
                    if 'ratio' in col.lower():
                        glesdmsaa_ratio_cols.append(col)
                    elif 'diff' in col.lower():
                        glesdmsaa_diff_cols.append(col)
            
            if glesdmsaa_ratio_cols or glesdmsaa_diff_cols:
                print(f"    Found glesdmsaa columns in imported data:")
                if glesdmsaa_ratio_cols:
                    print(f"      - Ratio columns: {', '.join(glesdmsaa_ratio_cols)}")
                if glesdmsaa_diff_cols:
                    print(f"      - Diff columns: {', '.join(glesdmsaa_diff_cols)}")
                
                # Create distribution tables from imported data
                imported_df = add_glesdmsaa_distribution_tables(imported_df, 
                                                                glesdmsaa_ratio_cols, 
                                                                glesdmsaa_diff_cols,
                                                                version_name,
                                                                version_summary_columns)
            
            return imported_df
        else:
            print(f"    ⚠️ Failed to import existing comparison page, generating new one")
    
    # ============================================================
    # CASE 2: Generate new comparison page
    # ============================================================
    print(f"    Generating new comparison page for version: {version_name}")
    
    # Find benchmarks that exist in ALL backends for this version
    backend_benchmarks = []
    for df in dataframes.values():
        backend_benchmarks.append(set(df['bench'].tolist()))
    
    if backend_benchmarks:
        # Find intersection of all benchmarks across all backends
        common_benchmarks = set.intersection(*backend_benchmarks)
        benches = sorted(list(common_benchmarks))
        print(f"    Filtered benchmarks: {len(common_benchmarks)} common out of {len(set.union(*backend_benchmarks))} total")
    else:
        benches = []
    
    if not benches:
        print(f"    WARNING: No common benchmarks found across all backends for version {version_name}")
        return pd.DataFrame()
    
    # Create ordered ID column
    ordered_ids = list(range(1, len(benches) + 1))
    
    # Prepare comparison data
    comparison_data = {
        'ID': ordered_ids,
        'Bench': benches
    }
    
    # Add mean columns for each backend (only for common benchmarks)
    backend_columns = {}  # Store column names for formula reference
    for col_name, df in sorted(dataframes.items()):
        mean_dict = dict(zip(df['bench'], df['mean']))
        
        # Get backend name from mapping (already extracted)
        display_name = backend_mapping.get(col_name, col_name)
        
        comparison_data[display_name] = [mean_dict.get(bench, float('nan')) for bench in benches]
        backend_columns[display_name] = display_name
    
    print(f"    Backend columns found: {', '.join(backend_columns.keys())}")
    
    # ============================================================
    # EXISTING LOGIC: Generate ratio_configs dynamically
    # ============================================================
    # 1. Compare any backend (except glesdmsaa) with glesdmsaa
    # 2. Compare any backend starting with 'gr' with grdawn_vk
    # 3. Compare grvk with vkdmsaa if both exist
    ratio_configs = []
    
    # Get all backends except glesdmsaa
    other_backends = [b for b in backend_columns.keys() if b != 'glesdmsaa']
    
    # 1. Compare each backend with glesdmsaa
    for backend in other_backends:
        ratio_configs.append((f"{backend} vs glesdmsaa", backend, 'glesdmsaa'))
    
    # 2. Compare any backend starting with 'gr' with grdawn_vk (if grdawn_vk exists)
    if 'grdawn_vk' in backend_columns:
        gr_backends = [b for b in backend_columns.keys() if b.startswith('gr') and b != 'grdawn_vk']
        for backend in gr_backends:
            ratio_configs.append((f"{backend} vs grdawn_vk", backend, 'grdawn_vk'))
    
    # 3. Also add grvk vs vkdmsaa if both exist (skip if already added by above rules)
    if 'grvk' in backend_columns and 'vkdmsaa' in backend_columns:
        # Check if this config already exists
        already_added = False
        for config_name, _, _ in ratio_configs:
            if config_name == 'grvk vs vkdmsaa':
                already_added = True
                break
        if not already_added:
            ratio_configs.append(('grvk vs vkdmsaa', 'grvk', 'vkdmsaa'))
    
    print(f"    Generated {len(ratio_configs)} ratio configurations:")
    for config in ratio_configs:
        print(f"      - {config[0]} = {config[1]}/{config[2]}")
    
    # Track all ratio and diff columns
    all_ratio_columns = []
    all_diff_columns = []
    
    # Track glesdmsaa-specific columns (configs where denominator is glesdmsaa)
    glesdmsaa_ratio_cols = []
    glesdmsaa_diff_cols = []
    
    for config_name, num_backend, den_backend in ratio_configs:
        if num_backend in backend_columns and den_backend in backend_columns:
            # Ratio column (division)
            ratio_col_name = f"{config_name} (ratio)"
            comparison_data[ratio_col_name] = [f"FORMULA:{num_backend}/{den_backend}"] * len(benches)
            all_ratio_columns.append(ratio_col_name)
            
            # Diff column (subtraction)
            diff_col_name = f"{config_name} (diff)"
            comparison_data[diff_col_name] = [f"FORMULA:{num_backend}-{den_backend}"] * len(benches)
            all_diff_columns.append(diff_col_name)
            
            print(f"      ✓ Added ratio column: '{ratio_col_name}' = {num_backend}/{den_backend}")
            print(f"      ✓ Added diff column: '{diff_col_name}' = {num_backend}-{den_backend}")
            
            # Track glesdmsaa-specific columns (where denominator is glesdmsaa)
            if den_backend == 'glesdmsaa':
                glesdmsaa_ratio_cols.append(ratio_col_name)
                glesdmsaa_diff_cols.append(diff_col_name)
        else:
            print(f"      ✗ WARNING: Backends not found for {config_name}")
            if num_backend not in backend_columns:
                print(f"        Missing numerator: {num_backend}")
            if den_backend not in backend_columns:
                print(f"        Missing denominator: {den_backend}")
    
    print(f"    Total columns added: {len(all_ratio_columns) + len(all_diff_columns)} (ratio + diff)")
    if glesdmsaa_ratio_cols:
        print(f"    Glesdmsaa ratio columns: {len(glesdmsaa_ratio_cols)}")
    if glesdmsaa_diff_cols:
        print(f"    Glesdmsaa diff columns: {len(glesdmsaa_diff_cols)}")
    
    # Add summary columns from this version's comparison page if provided
    if version_summary_columns:
        for col_name, col_values in version_summary_columns.items():
            if len(col_values) == len(benches):
                comparison_data[col_name] = col_values
                print(f"      ✓ Added summary column: '{col_name}'")
            else:
                print(f"      ✗ WARNING: Summary column '{col_name}' has {len(col_values)} values, expected {len(benches)}")
    
    # Create DataFrame
    df = pd.DataFrame(comparison_data)
    
    # ============================================================
    # Add glesdmsaa distribution tables (for generated data)
    # ============================================================
    df = add_glesdmsaa_distribution_tables(df, glesdmsaa_ratio_cols, glesdmsaa_diff_cols, version_name, version_summary_columns)
    
    # Add average row at the end
    if len(df) > 0:
        # Create a new row with 'AVERAGE' in the Bench column
        avg_row = {'ID': '', 'Bench': 'AVERAGE'}
        
        # Calculate average for each column (excluding non-numeric columns)
        for col in df.columns:
            if col not in ['ID', 'Bench']:
                # Try to convert to numeric, if fails, leave as empty string
                try:
                    # Check if column contains formula placeholders
                    if df[col].dtype == 'object' and len(df) > 0:
                        first_val = df[col].iloc[0]
                        if isinstance(first_val, str) and first_val.startswith('FORMULA:'):
                            # For formula columns, skip average (formulas can't be averaged as strings)
                            avg_row[col] = ''
                            continue
                    
                    # Convert column to numeric, coercing errors to NaN
                    numeric_values = pd.to_numeric(df[col], errors='coerce')
                    if not numeric_values.isna().all():
                        # Calculate mean, ignoring NaN
                        avg_value = numeric_values.mean()
                        avg_row[col] = avg_value
                    else:
                        avg_row[col] = ''
                except:
                    avg_row[col] = ''
        
        # Append the average row to the DataFrame
        df = pd.concat([df, pd.DataFrame([avg_row])], ignore_index=True)
    
    return df

def create_cross_version_page(version_groups, all_backends, summary_columns, missing_report, 
                               version_benchmarks, baseline_version, compare_versions, comparison_type):
    """Create cross-version comparison page with baseline as reference."""
    if len(version_groups) <= 1:
        return None
    
    # Find benchmarks that exist in ALL backends of ALL versions
    all_version_benchmarks = []
    for version_name, version_data in version_groups.items():
        backend_benchmarks = []
        for df in version_data['dataframes'].values():
            backend_benchmarks.append(set(df['bench'].tolist()))
        
        if backend_benchmarks:
            version_common = set.intersection(*backend_benchmarks)
            all_version_benchmarks.append(version_common)
    
    if all_version_benchmarks:
        common_benchmarks = set.intersection(*all_version_benchmarks)
        benches = sorted(list(common_benchmarks))
        print(f"    Cross-version common benchmarks: {len(common_benchmarks)} across all versions")
    else:
        benches = []
    
    if not benches:
        print(f"    WARNING: No common benchmarks found across all versions")
        return pd.DataFrame()
    
    # Create version mapping
    version_mapping = {}
    for full_tag in version_groups.keys():
        if comparison_type in ["same_skia_different_api", "same_skia_same_api"]:
            api_match = API_REGEX.search(full_tag)
            if api_match:
                compare_ver = f"{api_match.group(1).lower()}{api_match.group(2)}"
                version_mapping[full_tag] = compare_ver
            else:
                version_mapping[full_tag] = full_tag
        elif comparison_type == "same_api_different_skia":
            skia_match = SKIA_REGEX.search(full_tag)
            if skia_match:
                compare_ver = f"{skia_match.group(1).lower()}{skia_match.group(2)}"
                version_mapping[full_tag] = compare_ver
            else:
                version_mapping[full_tag] = full_tag
        else:
            version_mapping[full_tag] = full_tag
    
    # Prepare comparison data
    comparison_data = {
        'ID': list(range(1, len(benches) + 1)),
        'Bench': benches
    }
    
    # Track column names
    column_names = {}
    
    # Add mean columns
    for full_tag, version_data in version_groups.items():
        compare_ver = version_mapping[full_tag]
        backend_mapping = version_data.get('backend_mapping', {})
        
        for col_name, df in sorted(version_data['dataframes'].items()):
            backend = backend_mapping.get(col_name, col_name)
            mean_dict = dict(zip(df['bench'], df['mean']))
            display_name = f"{backend}_{compare_ver}"
            comparison_data[display_name] = [mean_dict.get(bench, float('nan')) for bench in benches]
            column_names[display_name] = display_name
    
    # Add ratio and diff columns
    backends = ['grdawn_vk', 'glesdmsaa', 'vkdmsaa', 'grvk']
    ratio_columns = []
    diff_columns = []
    
    for i, version1 in enumerate(compare_versions):
        for version2 in compare_versions[i+1:]:
            for backend in backends:
                col1_name = f"{backend}_{version1}"
                col2_name = f"{backend}_{version2}"
                
                if col1_name in column_names and col2_name in column_names:
                    config_name = f"{backend}_{version1} vs {backend}_{version2}"
                    ratio_col = f"{config_name} (ratio)"
                    diff_col = f"{config_name} (diff)"
                    
                    comparison_data[ratio_col] = [f"FORMULA:{col1_name}/{col2_name}"] * len(benches)
                    comparison_data[diff_col] = [f"FORMULA:{col1_name}-{col2_name}"] * len(benches)
                    
                    ratio_columns.append(ratio_col)
                    diff_columns.append(diff_col)
    
    # Add summary columns if provided
    if summary_columns:
        for col_name, col_values in summary_columns.items():
            if len(col_values) == len(benches):
                comparison_data[col_name] = col_values
    
    # Create DataFrame
    df = pd.DataFrame(comparison_data)
    
    # Build distribution table data with FREQUENCY formulas
    if ratio_columns or diff_columns:
        print(f"\n    Building distribution tables with FREQUENCY...")
        
        # Ratio distribution bins: (value, label)
        ratio_bins = [
            (0.1, '[0, 0.1]'),
            (0.2, '(0.1, 0.2]'),
            (0.5, '(0.2, 0.5]'),
            (1.0, '(0.5, 1.0]'),
            (5.0, '(1.0, 5.0]'),
            (10.0, '(5.0, 10.0]'),
            (float('inf'), '>10')
        ]
        
        # Diff distribution bins: (value, label)
        diff_bins = [
            (-10.0, '<-10'),
            (-5.0, '[-10, -5)'),
            (-1.0, '[-5, -1)'),
            (-0.5, '[-1, -0.5)'),
            (0.0, '[-0.5, 0)'),
            (1.0, '[0, 1]'),
            (float('inf'), '>1')
        ]
        
        # Build ratio distribution table
        if ratio_columns:
            # Only store bin labels and values - no formula placeholders for each cell
            ratio_dist_data = {
                'Bin': [threshold for threshold, _ in ratio_bins if threshold != float('inf')] + [''],
                'Label': [label for _, label in ratio_bins]
            }
            
            # Add empty columns for each ratio column (will be filled with FREQUENCY formula in first cell)
            for col in ratio_columns:
                ratio_dist_data[col] = [''] * len(ratio_bins)
            
            ratio_dist_df = pd.DataFrame(ratio_dist_data)
            df.attrs['ratio_distribution'] = ratio_dist_df
            df.attrs['ratio_bins'] = [threshold for threshold, _ in ratio_bins if threshold != float('inf')]
            df.attrs['ratio_columns'] = ratio_columns
            print(f"    ✓ Created ratio distribution table with {len(ratio_bins)} bins and {len(ratio_columns)} columns")
        
        # Build diff distribution table
        if diff_columns:
            diff_dist_data = {
                'Bin': [threshold for threshold, _ in diff_bins if threshold != float('inf')] + [''],
                'Label': [label for _, label in diff_bins]
            }
            
            for col in diff_columns:
                diff_dist_data[col] = [''] * len(diff_bins)
            
            diff_dist_df = pd.DataFrame(diff_dist_data)
            df.attrs['diff_distribution'] = diff_dist_df
            df.attrs['diff_bins'] = [threshold for threshold, _ in diff_bins if threshold != float('inf')]
            df.attrs['diff_columns'] = diff_columns
            print(f"    ✓ Created diff distribution table with {len(diff_bins)} bins and {len(diff_columns)} columns")
    
    return df

def add_frequency_formulas(dist_sheet, bin_values, data_columns, main_sheet_name, col_letters, data_start_row, data_end_row):
    """Add FREQUENCY formulas as array formulas using openpyxl.worksheet.formula.ArrayFormula."""
    # Column A: Bin values
    bin_col_idx = 1
    bin_col_letter = get_column_letter(bin_col_idx)
    
    # Write bin values in column A (starting from row 2)
    num_bins = len(bin_values)
    for row_idx, bin_val in enumerate(bin_values):
        cell = dist_sheet.cell(row=row_idx + 2, column=bin_col_idx)
        cell.value = bin_val
    
    # Bin range: A2:A{N+1}
    bin_range = f"{bin_col_letter}2:{bin_col_letter}{num_bins + 1}"
    
    for col_name in data_columns:
        if col_name in col_letters:
            # Find which column this is in the distribution sheet
            dist_col_idx = None
            for idx, header in enumerate(dist_sheet[1], 1):
                if header.value == col_name:
                    dist_col_idx = idx
                    break
            
            if dist_col_idx is not None:
                main_col_letter = col_letters[col_name]
                data_range = f"'{main_sheet_name}'!{main_col_letter}{data_start_row}:{main_col_letter}{data_end_row}"
                
                # FREQUENCY formula
                # FREQUENCY returns N+1 values where N = number of bins
                frequency_formula = f"=FREQUENCY({data_range},{bin_range})"
                
                # Determine the range for the array formula
                # FREQUENCY returns N+1 values, so we need N+1 rows
                start_row = 2
                end_row = num_bins + 2  # N+1 values + offset for header
                col_letter = get_column_letter(dist_col_idx)
                array_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                
                # Set the array formula on the FIRST cell of the range only
                # This prevents the @ symbol from appearing
                first_cell = f"{col_letter}{start_row}"
                dist_sheet[first_cell] = ArrayFormula(
                    ref=array_range,
                    text=frequency_formula
                )

def create_distribution_chart(dist_sheet, dist_df, chart_title, x_axis_title, y_axis_title, chart_position):
    """Create a beautified clustered column chart for the distribution table."""
    # Determine the data range
    num_rows = len(dist_df)
    num_cols = len(dist_df.columns)
    
    # Data starts at row 2 (row 1 is header), column 3 (C) is first data column
    # X-axis labels are in column 2 (B) - the 'Label' column
    x_labels_col = 2  # Column B
    start_data_col = 3  # Column C
    
    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 6  # Style 6
    chart.height = 14  # Height in inches
    chart.width = 22   # Width in inches
    chart.overlap = 0
    chart.gapWidth = 250
    
    chart.layout = Layout(
        manualLayout=ManualLayout()
    )
    chart.layout.manualLayout.x = 0.01
    chart.layout.manualLayout.y = 0.05
    chart.layout.manualLayout.w = 0.94
    chart.layout.manualLayout.h = 0.68
    
    # Set title with bold and size 18
    chart.title = chart_title
    # Use CharacterProperties for font styling
    rt = chart.title.tx.rich
    para = rt.p[0]
    run = para.r[0]
    run.rPr = CharacterProperties(
        sz=1800,
        b=True
    )
    
    # Set legend position (bottom, no overlay)
    chart.legend = Legend()
    chart.legend.position = 'b'  # Bottom
    chart.legend.overlay = False  # Don't overlay chart
        
    # Get the data range for the chart
    data_start_row = 2  # First data row
    data_end_row = num_rows + 1  # Last data row
    
    # Add all data series without titles
    for col_idx in range(start_data_col, num_cols + 1):
        values = Reference(dist_sheet, 
                          min_col=col_idx, 
                          min_row=data_start_row, 
                          max_row=data_end_row)
        chart.add_data(values, titles_from_data=False)
    
    # Get categories (x-axis labels) - use the Label column
    categories = Reference(dist_sheet, 
                          min_col=x_labels_col, 
                          min_row=data_start_row, 
                          max_row=data_end_row)
    
    # Set the categories for the chart
    chart.set_categories(categories)
    
    # Configure x-axis title (bold)
    chart.x_axis.title = x_axis_title
    xrt = chart.x_axis.title.tx.rich
    xpara = xrt.p[0]
    xrun = xpara.r[0]
    xrun.rPr = CharacterProperties(
        sz=1100,
        b=True
    )

    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.crosses = 'autoZero'
    chart.x_axis.spPr = GraphicalProperties(
        ln=LineProperties(solidFill="000000")
    )
    #chart.x_axis.txPr = RichText(
    #    p=[
    #        Paragraph(
    #            pPr=ParagraphProperties(
    #                defRPr=CharacterProperties(sz=900)
    #            )
    #        )
    #    ]
    #)
    
    # Configure y-axis title (bold)
    chart.y_axis.title = y_axis_title
    yrt = chart.y_axis.title.tx.rich
    ypara = yrt.p[0]
    yrun = ypara.r[0]
    yrun.rPr = CharacterProperties(
        sz=1100,
        b=True
    )
    chart.y_axis.delete = False
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.crosses = 'autoZero'
    chart.y_axis.scaling.min = 0  # Start at 0

    chart.y_axis.spPr = GraphicalProperties(
        ln=LineProperties(solidFill="000000")
    )
    
    # Set titles for each series with different colors
    # Color palette for different series
    
    colors = [
        '4472C4',  # Blue
        'C0504D',  # Red
        '9BBB59',  # Olive Green
        '8064A2',  # Purple
        '4BACC6',  # Aqua
        'F79646',  # Orange
        '7F7F7F',  # Gray
        '93C47D',  # Light Green
    ]
    
    for idx, col_name in enumerate(dist_df.columns[start_data_col-1:]):
        series = chart.series[idx]
        # Create SeriesLabel with the column name
        series.tx = SeriesLabel(v=col_name)
        
        # Set different color for each series
        color_idx = idx % len(colors)
        series.graphicalProperties.solidFill = colors[color_idx]
    
    # Add data labels - only show value
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showPercent = False
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.position = 'outEnd'  # Outside top
    
    # Set data label font (smaller, clean)
    for series in chart.series:
        if series.dLbls:
            series.dLbls.font = DrawingFont(sz=900)  # 9 pt
    
    # Add the chart to the sheet
    dist_sheet.add_chart(chart, chart_position)
    print(f"    ✓ Added chart at {chart_position} with title '{chart_title}'")

def write_dataframe_with_formulas(writer, sheet_name, df, baseline_version, add_average_row=True):
    """Write dataframe to Excel with proper Excel formulas, average row outside table."""
    if df.empty:
        print(f"    WARNING: DataFrame for '{sheet_name}' is empty, skipping")
        return
    
    # Check if dataframe already has an average row
    has_average_in_data = False
    avg_data = None
    
    if len(df) > 0 and 'Bench' in df.columns:
        last_val = df.iloc[-1]['Bench'] if len(df) > 0 else None
        if last_val == 'AVERAGE':
            has_average_in_data = True
            avg_data = df.iloc[-1].to_dict()
            df = df.iloc[:-1]
            print(f"    Found average row in dataframe - will preserve it")
    
    # Check if imported dataframe has average in attrs
    has_imported_average = df.attrs.get('has_average', False)
    imported_avg_data = df.attrs.get('average_row', None)
    
    if not has_average_in_data and has_imported_average and imported_avg_data is not None:
        has_average_in_data = True
        avg_data = imported_avg_data
        print(f"    Using imported average row from attrs")
    
    # Write dataframe without formulas
    df_for_write = df.copy()
    
    # Store formula info
    formula_info = {}
    for col in df_for_write.columns:
        if len(df_for_write) > 0:
            col_formulas = []
            for idx, val in enumerate(df_for_write[col]):
                if isinstance(val, str) and val.startswith('FORMULA:'):
                    col_formulas.append((idx, val[8:]))
            if col_formulas:
                formula_info[col] = col_formulas
                for idx, _ in col_formulas:
                    df_for_write[col].iloc[idx] = None
    
    # Write the main dataframe
    df_for_write.to_excel(writer, sheet_name=sheet_name, index=False)
    
    workbook = writer.book
    sheet = workbook[sheet_name]
    
    # Build column letter mapping
    col_letters = {}
    for idx, col_name in enumerate(df.columns, 1):
        col_letters[col_name] = get_column_letter(idx)
    
    last_data_row = len(df) + 1
    data_start_row = 2
    data_end_row = last_data_row - 1
    
    # Add formulas to main sheet
    for col_name, formulas in formula_info.items():
        if col_name in col_letters:
            col_idx = list(df.columns).index(col_name) + 1
            
            for row_idx_in_df, formula_expr in formulas:
                excel_row = row_idx_in_df + 2
                
                if '/' in formula_expr and not formula_expr.startswith('='):
                    parts = formula_expr.split('/')
                    if len(parts) == 2:
                        num_col_name = parts[0].strip()
                        den_col_name = parts[1].strip()
                        if num_col_name in col_letters and den_col_name in col_letters:
                            num_col_letter = col_letters[num_col_name]
                            den_col_letter = col_letters[den_col_name]
                            formula = f"={num_col_letter}{excel_row}/{den_col_letter}{excel_row}"
                            cell = sheet.cell(row=excel_row, column=col_idx)
                            cell.value = formula
                            cell.number_format = "0.000"
                elif '-' in formula_expr and not formula_expr.startswith('='):
                    parts = formula_expr.split('-')
                    if len(parts) == 2:
                        num_col_name = parts[0].strip()
                        den_col_name = parts[1].strip()
                        if num_col_name in col_letters and den_col_name in col_letters:
                            num_col_letter = col_letters[num_col_name]
                            den_col_letter = col_letters[den_col_name]
                            formula = f"={num_col_letter}{excel_row}-{den_col_letter}{excel_row}"
                            cell = sheet.cell(row=excel_row, column=col_idx)
                            cell.value = formula
                            cell.number_format = "0.000"
                elif formula_expr.startswith('='):
                    cell = sheet.cell(row=excel_row, column=col_idx)
                    cell.value = formula_expr
                    if 'ratio' in col_name.lower() or 'diff' in col_name.lower():
                        cell.number_format = "0.000"
    
    # Add average row
    if add_average_row:
        if not has_average_in_data or avg_data is None:
            print(f"    No existing average row - calculating new one")
            avg_data = {}
            for col_name in df.columns:
                if col_name not in ['ID', 'Bench']:
                    should_have_average = False
                    
                    if col_name in formula_info:
                        should_have_average = True
                    else:
                        for val in df[col_name].head(5):
                            if val is not None:
                                try:
                                    float(val)
                                    should_have_average = True
                                    break
                                except (ValueError, TypeError):
                                    pass
                    
                    if should_have_average:
                        numeric_values = pd.to_numeric(df[col_name], errors='coerce')
                        if not numeric_values.isna().all():
                            avg_data[col_name] = numeric_values.mean()
                        else:
                            avg_data[col_name] = ''
                    else:
                        avg_data[col_name] = ''
            avg_data['ID'] = ''
            avg_data['Bench'] = 'AVERAGE'
            has_average_in_data = True
            print(f"    ✓ Calculated new average row")
        
        if has_average_in_data and avg_data:
            avg_row_excel = last_data_row + 2
            
            for idx, col_name in enumerate(df.columns, 1):
                if col_name == 'Bench':
                    cell = sheet.cell(row=avg_row_excel, column=idx)
                    cell.value = "AVERAGE"
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_name not in ['ID']:
                    should_have_average = False
                    
                    if col_name in formula_info:
                        should_have_average = True
                    else:
                        for check_row in range(2, min(last_data_row + 1, 10)):
                            cell_value = sheet.cell(row=check_row, column=idx).value
                            if cell_value is not None and isinstance(cell_value, (int, float)):
                                should_have_average = True
                                break
                            elif cell_value is not None and isinstance(cell_value, str):
                                try:
                                    float(cell_value)
                                    should_have_average = True
                                    break
                                except:
                                    pass
                    
                    if should_have_average:
                        if col_name in avg_data and avg_data[col_name] not in ['', None]:
                            cell = sheet.cell(row=avg_row_excel, column=idx)
                            cell.value = avg_data[col_name]
                            cell.number_format = "0.000"
                        else:
                            col_letter = get_column_letter(idx)
                            avg_formula = f"=AVERAGE({col_letter}2:{col_letter}{last_data_row + 1})"
                            cell = sheet.cell(row=avg_row_excel, column=idx)
                            cell.value = avg_formula
                            cell.number_format = "0.000"
                        
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell = sheet.cell(row=avg_row_excel, column=idx)
                        cell.value = ""
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"    ✓ Added average row at row {avg_row_excel} (outside table)")
    
    # Write distribution tables with FREQUENCY using ArrayFormula
    main_sheet_name = sheet_name
    main_workbook = writer.book
    
    def format_distribution_sheet(dist_sheet, dist_df):
        """Format distribution sheet with header styling."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in dist_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in dist_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            dist_sheet.column_dimensions[column_letter].width = adjusted_width
        
        dist_sheet.freeze_panes = dist_sheet['A2']
      
    # Write ratio distribution table
    if df.attrs.get('ratio_distribution') is not None:
        ratio_dist_df = df.attrs['ratio_distribution']
        ratio_bins = df.attrs.get('ratio_bins', [])
        ratio_columns = df.attrs.get('ratio_columns', [])
        dist_sheet_name = sheet_name.replace("_comparison","_ratio_dist")
        
        ratio_dist_df.to_excel(writer, sheet_name=dist_sheet_name, index=False)
        dist_sheet = workbook[dist_sheet_name]
        
        add_frequency_formulas(dist_sheet, ratio_bins, ratio_columns, main_sheet_name, col_letters, data_start_row, data_end_row)
        format_distribution_sheet(dist_sheet, ratio_dist_df)
        print(f"    ✓ Wrote ratio distribution table to sheet '{dist_sheet_name}' with FREQUENCY (ArrayFormula)")
        
        # Create chart for ratio distribution
        # Extract Skia version from sheet name or use default
        chart_title = f"{baseline_version} Nanobench Time Ratio Distribution by Range()"
        x_title = "Time Ratio Range"
        y_title = "Count"
        
        # Position chart below the data (starting at row after data + 2)
        chart_start_row = len(ratio_dist_df) + 5
        chart_position = f"C{chart_start_row}"
        create_distribution_chart(dist_sheet, ratio_dist_df, chart_title, x_title, y_title, chart_position)
    
    # Write diff distribution table
    if df.attrs.get('diff_distribution') is not None:
        diff_dist_df = df.attrs['diff_distribution']
        diff_bins = df.attrs.get('diff_bins', [])
        diff_columns = df.attrs.get('diff_columns', [])
        dist_sheet_name = sheet_name.replace("_comparison", "_diff_dist")
        
        diff_dist_df.to_excel(writer, sheet_name=dist_sheet_name, index=False)
        dist_sheet = workbook[dist_sheet_name]
        
        add_frequency_formulas(dist_sheet, diff_bins, diff_columns, main_sheet_name, col_letters, data_start_row, data_end_row)
        format_distribution_sheet(dist_sheet, diff_dist_df)
        print(f"    ✓ Wrote diff distribution table to sheet '{dist_sheet_name}' with FREQUENCY (ArrayFormula)")
        
        # Create chart for diff distribution
        chart_title = f"{baseline_version} Nanobench Time Diff Distribution by Range()"
        x_title = "Time Diff Range(ms)"
        y_title = "Count"
        
        # Position chart below the data (starting at row after data + 2)
        chart_start_row = len(diff_dist_df) + 5
        chart_position = f"C{chart_start_row}"
        create_distribution_chart(dist_sheet, diff_dist_df, chart_title, x_title, y_title, chart_position)
    
    # ============================================================
    # Write glesdmsaa ratio distribution table
    # ============================================================
    if df.attrs.get('glesdmsaa_ratio_distribution') is not None:
        glesdmsaa_ratio_dist_df = df.attrs['glesdmsaa_ratio_distribution']
        glesdmsaa_ratio_bins = df.attrs.get('glesdmsaa_ratio_bins', [])
        glesdmsaa_ratio_columns = df.attrs.get('glesdmsaa_ratio_columns', [])
        dist_sheet_name = sheet_name.replace("_comparison", "_ratio_dist")

        glesdmsaa_ratio_dist_df.to_excel(writer, sheet_name=dist_sheet_name, index=False)
        dist_sheet = workbook[dist_sheet_name]

        add_frequency_formulas(dist_sheet, glesdmsaa_ratio_bins, glesdmsaa_ratio_columns,
                              main_sheet_name, col_letters, data_start_row, data_end_row)
        format_distribution_sheet(dist_sheet, glesdmsaa_ratio_dist_df)
        print(f"    ✓ Wrote glesdmsaa ratio distribution table to sheet '{dist_sheet_name}' with FREQUENCY (ArrayFormula)")

        # Create chart for glesdmsaa ratio distribution
        compare_prefix = sheet_name.replace("_comparison","")
        chart_title = f"{baseline_version} {compare_prefix} Nanobench Time Ratio Distribution (vs glesdmsaa)"
        x_title = "Time Ratio Range"
        y_title = "Count"

        chart_start_row = len(glesdmsaa_ratio_dist_df) + 5
        chart_position = f"C{chart_start_row}"
        create_distribution_chart(dist_sheet, glesdmsaa_ratio_dist_df, chart_title, x_title, y_title, chart_position)

    # ============================================================
    # Write glesdmsaa diff distribution table
    # ============================================================
    if df.attrs.get('glesdmsaa_diff_distribution') is not None:
        glesdmsaa_diff_dist_df = df.attrs['glesdmsaa_diff_distribution']
        glesdmsaa_diff_bins = df.attrs.get('glesdmsaa_diff_bins', [])
        glesdmsaa_diff_columns = df.attrs.get('glesdmsaa_diff_columns', [])
        dist_sheet_name = sheet_name.replace("_comparison", "_diff_dist")

        glesdmsaa_diff_dist_df.to_excel(writer, sheet_name=dist_sheet_name, index=False)
        dist_sheet = workbook[dist_sheet_name]

        add_frequency_formulas(dist_sheet, glesdmsaa_diff_bins, glesdmsaa_diff_columns,
                              main_sheet_name, col_letters, data_start_row, data_end_row)
        format_distribution_sheet(dist_sheet, glesdmsaa_diff_dist_df)
        print(f"    ✓ Wrote glesdmsaa diff distribution table to sheet '{dist_sheet_name}' with FREQUENCY (ArrayFormula)")

        # Create chart for glesdmsaa diff distribution
        compare_prefix = sheet_name.replace("_comparison","")
        chart_title = f"{baseline_version} {compare_prefix} Nanobench Time Diff Distribution (vs glesdmsaa)"
        x_title = "Time Diff Range (ms)"
        y_title = "Count"

        chart_start_row = len(glesdmsaa_diff_dist_df) + 5
        chart_position = f"C{chart_start_row}"
        create_distribution_chart(dist_sheet, glesdmsaa_diff_dist_df, chart_title, x_title, y_title, chart_position)

    # ============================================================
    # Write draw call distribution tables using existing helpers
    # ============================================================
    # Find all draw distribution attributes
    ratio_attrs = [key for key in df.attrs.keys() if key.endswith('_ratio_draw_distribution')]
    diff_attrs = [key for key in df.attrs.keys() if key.endswith('_diff_draw_distribution')]
    
    # Write ratio draw distribution tables
    for attr_name in ratio_attrs:
        summary_col = attr_name.replace('_ratio_draw_distribution', '')
        dist_df = df.attrs[attr_name]
        
        if dist_df is not None and not dist_df.empty:
            dist_sheet_name = sheet_name.replace("_comparison",f"_{summary_col}_draw_dist_ratio")
            
            # Write the dataframe
            dist_df.to_excel(writer, sheet_name=dist_sheet_name, index=False)
            dist_sheet = workbook[dist_sheet_name]
            
            # Reuse existing formatting helper
            format_distribution_sheet(dist_sheet, dist_df)
            
            print(f"    ✓ Wrote ratio draw distribution table for summary column '{summary_col}' to sheet '{dist_sheet_name}'")
            
            # Create chart for ratio draw distribution
            compare_prefix = sheet_name.replace("_comparison","")
            chart_title = f"{baseline_version} {compare_prefix} Nanobench {summary_col} Draw Call Distribution By Ratio (grdawn_vk vs glesdmsaa)"
            x_title = "Time Ratio Range"
            y_title = "Weighted Count"
            chart_start_row = len(dist_df) + 5
            chart_position = f"C{chart_start_row}"
            
            # Reuse existing chart helper (it works the same way - categories from column B, data from columns C+)
            create_distribution_chart(dist_sheet, dist_df, chart_title, x_title, y_title, chart_position)
    
    # Write diff draw distribution tables
    for attr_name in diff_attrs:
        summary_col = attr_name.replace('_diff_draw_distribution', '')
        dist_df = df.attrs[attr_name]
        
        if dist_df is not None and not dist_df.empty:
            dist_sheet_name = sheet_name.replace("_comparison", f"_{summary_col}_draw_dist_diff")
            
            # Write the dataframe
            dist_df.to_excel(writer, sheet_name=dist_sheet_name, index=False)
            dist_sheet = workbook[dist_sheet_name]
            
            # Reuse existing formatting helper
            format_distribution_sheet(dist_sheet, dist_df)
            
            print(f"    ✓ Wrote diff draw distribution table for summary column '{summary_col}' to sheet '{dist_sheet_name}'")
            
            # Create chart for diff draw distribution
            compare_prefix = sheet_name.replace("_comparison","")
            chart_title = f"{baseline_version} {compare_prefix} Nanobench {summary_col} Draw Call Distribution By Diff(grdawn_vk vs glesdmsaa)"
            x_title = "Time Diff Range(ms)"
            y_title = "Weighted Count"
            chart_start_row = len(dist_df) + 5
            chart_position = f"C{chart_start_row}"
            
            # Reuse existing chart helper
            create_distribution_chart(dist_sheet, dist_df, chart_title, x_title, y_title, chart_position)
    
    return df

def apply_table_formatting_to_sheet(sheet, df):
    """Apply Excel table formatting to a sheet."""
    
    # Determine the range of the table (exclude average row)
    start_row = 1
    start_col = 1
    end_row = len(df) + 1  # +1 for header
    end_col = len(df.columns)
    
    # Create table range reference
    table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    
    # Create table
    table_name = f"Table_{sheet.title.replace(' ', '_')[:20]}"
    
    # Remove existing table if it exists
    for existing_table in list(sheet.tables.keys()):
        if existing_table.startswith("Table_"):
            del sheet.tables[existing_table]
    
    table = Table(displayName=table_name, ref=table_range)
    
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    sheet.add_table(table)
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze panes
    sheet.freeze_panes = sheet['B2']
    
    print(f"    ✓ Applied table formatting to range: {table_range}")
    
    return table

def backup_original_sheets(writer, all_original_sheets):
    """Backup original backend sheets into the output workbook."""
    workbook = writer.book
    
    print(f"\n📋 Backing up original backend pages...")
    
    for sheet_name, df in all_original_sheets.items():
        # Clean sheet name (Excel has 31 char limit)
        clean_sheet_name = sheet_name[:31]
        
        # Check if sheet already exists and rename if needed
        final_sheet_name = clean_sheet_name
        counter = 1
        while final_sheet_name in workbook.sheetnames:
            final_sheet_name = f"{clean_sheet_name[:27]}_{counter}"
            counter += 1
        
        # Write dataframe to sheet
        df.to_excel(writer, sheet_name=final_sheet_name, index=False)
        
        # Get the sheet and format it
        sheet = workbook[final_sheet_name]
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Freeze header row
        sheet.freeze_panes = sheet['A2']
        
        print(f"  ✓ Backed up '{sheet_name}' -> sheet '{final_sheet_name}' ({len(df)} rows)")

def print_summary(version_groups, 
                  output_file, version_num, all_original_sheets, summary_columns, 
                  missing_report, unique_benchmarks, duplicate_report, 
                  missing_backends, single_version_backends, baseline_version, 
                  compare_versions, comparison_type):
    """Print a summary of the analysis."""
    print("\n" + "="*60)
    print("✅ Analysis complete!")
    
    print(f"\n📊 Versions processed: {len(version_groups)}")
    for version in version_groups.keys():
        num_backends = len(version_groups[version]['dataframes'])
        print(f"  - {version}: {num_backends} backends")
    
    print(f"\n🎯 Baseline version: {baseline_version}")
    if compare_versions:
        print(f"   Compare versions: {', '.join(compare_versions)}")
    
    # Print comparison type
    comparison_type_names = {
        "same_skia_different_api": "Same Skia, Different API",
        "same_api_different_skia": "Same API, Different Skia",
        "same_skia_same_api": "Same Skia and Same API",
        "mixed_versions": "Mixed Versions",
        "single": "Single Version"
    }
    print(f"\n📋 Comparison type: {comparison_type_names.get(comparison_type, comparison_type)}")
    
    # Print column/backend coverage summary
    if missing_backends:
        print(f"\n⚠️  Backends missing from some versions: {len(missing_backends)}")
        for backend, info in missing_backends.items():
            print(f"    - '{backend}' missing in: {', '.join(info['missing_in'])}")
    
    if single_version_backends:
        print(f"\n📌 Backends unique to single version: {len(single_version_backends)}")
    
    # Print duplicate benchmark summary
    if duplicate_report:
        total_duplicates = sum(len(sheets) for sheets in duplicate_report.values())
        print(f"\n⚠️  Duplicate Benchmarks Found: {total_duplicates} sheets with duplicates")
    
    # Print missing benchmark summary
    if missing_report:
        total_missing = sum(len(m) for m in missing_report.values())
        print(f"\n⚠️  Missing Benchmarks: {total_missing} total missing entries across versions")
    
    if unique_benchmarks:
        print(f"\n📌 Benchmarks unique to single version: {len(unique_benchmarks)}")
    
    print(f"\n📋 Original sheets backed up: {len(all_original_sheets)}")
    
    if summary_columns:
        print(f"\n📋 Summary columns extracted from existing comparison: {len(summary_columns)}")
    
    print(f"\n📁 Output file: {output_file}")
    print(f"   Version number: v{version_num}")
    
    print("\n📑 Sheets in output workbook:")
    for version in compare_versions:
        print(f"  - {version}_comparison (compare version page)")
    
    # Only show cross-version sheet if it was created
    if len(version_groups) > 1 and comparison_type != "same_skia_same_api":
        print("  - cross_version_comparison (baseline + compare versions side-by-side)")
    elif comparison_type == "same_skia_same_api":
        print("  - ℹ️ No cross-version comparison (all versions have same Skia and API)")
    elif len(version_groups) <= 1:
        print("  - ℹ️ No cross-version comparison (only one version found)")
    
    print("  - backend_version (original data backups - unfiltered)")
    
    print("\n💡 Tips for using the Excel file:")
    print("  1. Compare version pages show benchmarks that exist in ALL backends of that version")
    if len(version_groups) > 1 and comparison_type != "same_skia_same_api":
        print("  2. Cross-version page shows baseline vs compare versions side-by-side")
    print("  3. Original backup sheets contain complete unfiltered data")
    print("  4. Use drop-down arrows in headers to sort/filter data")
    print("  5. First row and column are frozen for easy scrolling")

def main():
    """Main function to orchestrate the script."""
    print("="*60)
    print(f"Benchmark Analysis Tool - Version {VERSION}")
    print("="*60)
    
    # Validate arguments
    excel_files = validate_arguments()
    
    print(f"\n📄 Excel files ({len(excel_files)}):")
    for excel_file in excel_files:
        print(f"  - {excel_file}")
    
    # Read all Excel files
    (all_dataframes, all_backends, version_info, version_groups, 
     all_original_sheets, primary_summary_columns, missing_report, 
     unique_benchmarks, duplicate_report, missing_backends, 
     single_version_backends, version_benchmarks, baseline_version, 
     compare_versions, comparison_type, summary_columns_by_version, 
     has_comparison_page) = read_multiple_excel_files(excel_files)
    
    # Get unique benches
    all_benches = set()
    for df in all_dataframes.values():
        all_benches.update(df['bench'].tolist())
    print(f"\n📋 Found {len(all_benches)} unique benchmarks")
    
    # Generate output filename
    output_file, version_num = generate_output_filename(baseline_version)
    
    # Write to Excel
    print(f"\n💾 Generating Excel workbook: {output_file} (version {version_num})")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Create comparison pages for each compare version
        for compare_version in compare_versions:
            full_version = None
            for v in version_groups.keys():
                if compare_version in v:
                    full_version = v
                    break
            
            if full_version and full_version in version_groups:
                print(f"\n  Creating comparison page for version: {compare_version}")
                version_data = version_groups[full_version]
                missing_for_version = missing_report.get(full_version, set())
                
                version_summary = version_data.get('summary_columns', primary_summary_columns)
                
                version_df = create_version_comparison_page(
                    version_data, full_version, 
                    version_summary, missing_for_version, version_benchmarks,
                    baseline_version, is_baseline=False
                )
                
                if not version_df.empty:
                    sheet_name = f"{compare_version}_comparison"[:31]
                    write_dataframe_with_formulas(writer, sheet_name, version_df, baseline_version)
                    
                    workbook = writer.book
                    sheet = workbook[sheet_name]
                    apply_table_formatting_to_sheet(sheet, version_df)
                    print(f"    ✓ Created '{sheet_name}' with {len(version_df)} benchmarks, {len(version_df.columns)} columns")
                else:
                    print(f"    ✗ Skipping '{compare_version}_comparison' - no common benchmarks found")
        
        # Create cross-version comparison page
        if len(version_groups) > 1 and comparison_type != "same_skia_same_api":
            print(f"\n  Creating cross-version comparison page")
            cross_version_df = create_cross_version_page(
                version_groups, all_backends, primary_summary_columns, missing_report, 
                version_benchmarks, baseline_version, compare_versions, comparison_type
            )
            if cross_version_df is not None and not cross_version_df.empty:
                # Write the main cross-version comparison
                cross_sheet_name = f"cross_version_comparison"
                write_dataframe_with_formulas(writer, cross_sheet_name, cross_version_df, baseline_version)
                
                workbook = writer.book
                sheet = workbook[cross_sheet_name]
                apply_table_formatting_to_sheet(sheet, cross_version_df)
                print(f"    ✓ Created '{cross_sheet_name}' with {len(cross_version_df)} benchmarks, {len(cross_version_df.columns)} columns")
                
                # Distribution tables are automatically written by write_dataframe_with_formulas
            else:
                print(f"    ✗ Skipping 'cross_version_comparison' - no common benchmarks across all versions")
        else:
            if comparison_type == "same_skia_same_api":
                print(f"\n  ℹ️ Skipping cross-version comparison - all versions have same Skia and API")
            elif len(version_groups) <= 1:
                print(f"\n  ℹ️ Skipping cross-version comparison - only one version found")
        
        # Backup original backend sheets
        backup_original_sheets(writer, all_original_sheets)
    
    # Print summary
    print_summary(version_groups, 
                  output_file, version_num, all_original_sheets, primary_summary_columns, 
                  missing_report, unique_benchmarks, duplicate_report, 
                  missing_backends, single_version_backends, baseline_version, 
                  compare_versions, comparison_type)
    
    print("\n" + "="*60)
    print(f"Benchmark Analysis Tool v{VERSION} - Execution Complete")
    print("="*60)

if __name__ == "__main__":
    main()