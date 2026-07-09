#!/usr/bin/env python3
"""
Script to process CSV benchmark files and generate an Excel workbook with comparison.
Usage: python script.py <folder_path1> [folder_path2] <csv_file1> [csv_file2 ...]
Maximum 2 folder paths can be provided as the first parameters.
If a folder parameter is not a valid directory, trace parser logic will be skipped for that folder.
"""

import sys
import os
import json
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings
warnings.filterwarnings('ignore')

def validate_arguments():
    """Validate command line arguments."""
    if len(sys.argv) < 3:
        print("Error: At least 3 arguments required (at least one folder and one CSV file)")
        print(f"Usage: {sys.argv[0]} <folder_path1> [folder_path2] <csv_file1> [csv_file2 ...]")
        print("Note: Maximum 2 folder paths can be provided as the first parameters")
        sys.exit(1)
    
    # Parse arguments: first 1-2 could be folders, rest are CSV files
    folder_paths = []
    csv_files = []
    
    # Check each argument to determine if it's a folder or CSV
    for arg in sys.argv[1:]:
        if arg.lower().endswith('.csv'):
            csv_files.append(arg)
        elif len(folder_paths) < 2:  # Max 2 folders
            folder_paths.append(arg)
        else:
            # If we already have 2 folders and this isn't a CSV, it's an error
            print(f"Error: Maximum 2 folder parameters allowed. Extra parameter: '{arg}'")
            sys.exit(1)
    
    if len(csv_files) == 0:
        print("Error: At least one CSV file must be provided")
        sys.exit(1)
    
    # Check if folders exist (warn but don't exit)
    folder_exists = []
    for folder_path in folder_paths:
        exists = os.path.isdir(folder_path)
        folder_exists.append(exists)
        if not exists:
            print(f"Warning: Folder path '{folder_path}' does not exist - trace analysis will be skipped for this folder")
    
    # Check CSV files exist
    for csv_file in csv_files:
        if not os.path.isfile(csv_file):
            print(f"Error: CSV file '{csv_file}' does not exist")
            sys.exit(1)
    
    return folder_paths, csv_files, folder_exists

def detect_backend_type(folder_path):
    """Detect backend type from folder name."""
    folder_name = Path(folder_path).name.lower()
    if 'gr' in folder_name:
        return 'graphite'
    else:
        return 'ganesh'

def sanitize_bench_name_for_file(bench_name):
    """Remove dots from bench name for file lookup."""
    # Remove all dots from the benchmark name
    sanitized = bench_name.replace('.', '')
    return sanitized

def handle_duplicate_benchmarks(df, csv_file):
    """Detect and handle duplicate benchmark names in a dataframe.
    
    Returns:
        tuple: (processed_df, warning_messages)
    """
    warnings_list = []
    
    # Check for duplicates in 'bench' column
    duplicate_benches = df[df['bench'].duplicated(keep=False)]
    
    if not duplicate_benches.empty:
        file_name = Path(csv_file).name
        print(f"\n⚠️  WARNING: Duplicate benchmark names found in {file_name}")
        
        # Get unique duplicate bench names and their counts
        duplicate_counts = df['bench'].value_counts()
        duplicates = duplicate_counts[duplicate_counts > 1]
        
        for bench_name, count in duplicates.items():
            print(f"   - '{bench_name}' appears {count} times")
            
            # Get all rows for this duplicate benchmark
            duplicate_rows = df[df['bench'] == bench_name]
            warnings_list.append(f"Benchmark '{bench_name}' appears {count} times in {file_name}")
            
            # Show the mean values for each duplicate
            for idx, row in duplicate_rows.iterrows():
                print(f"     Row {idx + 2}: mean={row['mean']:.4f}" if 'mean' in row else f"     Row {idx + 2}: {row.to_dict()}")
        
        # Strategy: Keep the first occurrence and warn about others
        print(f"\n   📌 Strategy: Keeping first occurrence, dropping {len(duplicate_benches) - len(duplicates)} duplicate rows")
        df = df.drop_duplicates(subset=['bench'], keep='first')
        
    return df, warnings_list

def read_csv_files(csv_files):
    """Read CSV files, validate columns, and detect duplicate benchmarks."""
    dataframes = {}
    all_warnings = []
    
    for csv_file in csv_files:
        try:
            df = pd.read_csv(csv_file)
            
            # Check required columns
            if 'bench' not in df.columns or 'mean' not in df.columns:
                print(f"Error: CSV file '{csv_file}' must contain 'bench' and 'mean' columns")
                sys.exit(1)
            
            # Check for and handle duplicate benchmarks
            df, warnings = handle_duplicate_benchmarks(df, csv_file)
            if warnings:
                all_warnings.extend(warnings)
            
            dataframes[csv_file] = df
            
        except Exception as e:
            print(f"Error reading CSV file '{csv_file}': {e}")
            sys.exit(1)
    
    # Print summary of all duplicate warnings
    if all_warnings:
        print("\n" + "="*60)
        print("📋 DUPLICATE BENCHMARK SUMMARY:")
        for warning in all_warnings:
            print(f"  • {warning}")
        print("="*60 + "\n")
    
    return dataframes

def analyze_ftrace_files_ganesh(folder_path, benches, folder_name):
    """Analyze ftrace JSON files for Ganesh backend."""
    folder = Path(folder_path)
    draw_types_map = {}
    
    # Pattern to match SurfaceDrawContext::draw* functions
    draw_pattern = re.compile(r'SurfaceDrawContext::draw([a-zA-Z]+)')
    # Pattern to match GrDrawingManager::flush specifically
    flush_pattern = re.compile(r'GrDrawingManager::flush')
    
    for bench in benches:
        # Sanitize bench name by removing dots for file lookup
        sanitized_bench = sanitize_bench_name_for_file(bench)
        json_file = folder / f"{sanitized_bench}.json"
        
        # Also try original name if sanitized version doesn't exist
        if not json_file.exists():
            json_file = folder / f"{bench}.json"
        
        draw_counts = {}
        flush_count = 0
        
        if json_file.exists():
            try:
                with open(json_file, 'r') as f:
                    data = json.load(f)
                
                # Handle different ftrace JSON structures
                def process_trace(trace_data):
                    nonlocal flush_count
                    if isinstance(trace_data, dict):
                        # Check for function name in common fields
                        func_name = trace_data.get('func') or trace_data.get('function') or trace_data.get('name')
                        if func_name:
                            # Count draw types
                            draw_match = draw_pattern.search(func_name)
                            if draw_match:
                                draw_type = draw_match.group(1)
                                draw_counts[draw_type] = draw_counts.get(draw_type, 0) + 1
                            
                            # Count GrDrawingManager::flush calls
                            if flush_pattern.search(func_name):
                                flush_count += 1
                        
                        # Recursively process nested structures
                        for value in trace_data.values():
                            if isinstance(value, (dict, list)):
                                process_trace(value)
                    elif isinstance(trace_data, list):
                        for item in trace_data:
                            process_trace(item)
                
                process_trace(data)
                
                if draw_counts and flush_count > 0:
                    # Calculate division result for each draw type: total_count / flush_count
                    sorted_counts = sorted(draw_counts.items(), key=lambda x: x[1], reverse=True)
                    result_parts = []
                    for draw_type, count in sorted_counts:
                        division_result = count / flush_count
                        # Format with 2 decimal places for the division result
                        result_parts.append(f"{draw_type}:{division_result:.2f}")
                    draw_types_map[bench] = ', '.join(result_parts)
                elif draw_counts:
                    draw_types_map[bench] = f"No flush calls found. Draw types: {', '.join([f'{dt}:{cnt}' for dt, cnt in draw_counts.items()])}"
                else:
                    draw_types_map[bench] = "No draw functions found"
                    
            except Exception as e:
                draw_types_map[bench] = f"Error parsing JSON: {str(e)[:50]}"
        else:
            # Check if we tried sanitized version but original might have existed
            if sanitized_bench != bench:
                draw_types_map[bench] = f"JSON file not found (tried: {sanitized_bench}.json and {bench}.json)"
            else:
                draw_types_map[bench] = f"JSON file not found: {json_file}"
    
    return draw_types_map

def analyze_ftrace_files_graphite(folder_path, benches):
    """Analyze ftrace JSON files for Graphite backend."""
    folder = Path(folder_path)
    draw_types_map = {}
    
    # Compile regex patterns
    submit_pattern = re.compile(r'skiatest::graphite::GraphiteTestContext::submitRecordingAndWaitOnSync')
    snap_pattern = re.compile(r'skgpu::graphite::Recorder::snap')
    draw_pass_pattern = re.compile(r'skgpu::graphite::DrawList::snapDrawPass')
    
    for bench in benches:
        sanitized_bench = sanitize_bench_name_for_file(bench)
        json_file = folder / f"{sanitized_bench}.json"
        submissions = []
        
        if json_file.exists():
            try:
                with open(json_file, 'r') as f:
                    data = json.load(f)
                
                state = [None, []]
                
                def process_trace(trace_data):
                    if isinstance(trace_data, dict):
                        func_name = trace_data.get('func') or trace_data.get('function') or trace_data.get('name')
                        
                        if func_name:
                            # Start of submission
                            if submit_pattern.search(func_name):
                                state[0] = {
                                    'start_time': trace_data.get('timestamp', 0),
                                    'renderers': [],
                                    'draw_count': 0
                                }
                                state[1] = []
                            
                            # Snap event - update end_time and snap_time
                            elif snap_pattern.search(func_name) and state[0] is not None:
                                state[0]['snap_time'] = trace_data.get('timestamp', 0)
                                state[0]['end_time'] = trace_data.get('timestamp', 0)
                            
                            # Draw pass event - parse draw count only
                            elif draw_pass_pattern.search(func_name) and state[0] is not None:
                                draw_count = 0
                                args = trace_data.get('args', {})
                                for key, value in args.items():
                                    if 'draw count' in key.lower() or 'draw_count' in key.lower():
                                        try:
                                            draw_count = int(value)
                                            break
                                        except:
                                            pass
                                
                                state[0]['draw_count'] = draw_count
                                
                                if state[1]:
                                    state[0]['renderers'] = state[1]
                                
                                submissions.append(state[0])
                                state[0] = None
                                state[1] = []
                            
                            # Check for renderer in event arguments
                            args = trace_data.get('args', {})
                            for key, value in args.items():
                                if key.lower().startswith('renderer'):
                                    renderer_name = str(value)
                                    if state[0] is not None:
                                        state[1].append(renderer_name)
                                    break
                        
                        for key, value in trace_data.items():
                            if isinstance(value, (dict, list)):
                                process_trace(value)
                    elif isinstance(trace_data, list):
                        for item in trace_data:
                            process_trace(item)
                
                process_trace(data)
                
                if submissions:
                    # Group submissions by config
                    config_groups = {}
                    
                    for sub in submissions:
                        renderer_count = len(sub['renderers'])
                        draw_count = sub.get('draw_count', 0)
                        config_key = f"m{renderer_count}_d{draw_count}"
                        
                        if config_key not in config_groups:
                            config_groups[config_key] = {
                                'count': 0,
                                'renderer_count': renderer_count,
                                'draw_count': draw_count,
                                'renderers': sub['renderers'][:] if sub['renderers'] else []
                            }
                        config_groups[config_key]['count'] += 1
                    
                    # Sort configs by count descending
                    sorted_configs = sorted(config_groups.items(), key=lambda x: x[1]['count'], reverse=True)
                    
                    # Build summary parts
                    summary_parts = []
                    
                    # Format ALL configs as sub#id: N[summary]
                    for idx, (config_key, data) in enumerate(sorted_configs):
                        renderer_count = data['renderer_count']
                        draw_count = data['draw_count']
                        count = data['count']
                        renderers = data['renderers']
                        
                        # Check for error condition: rdr < draw
                        if renderer_count < draw_count:
                            # Log error message to terminal
                            print(f"    ERROR in {bench}: renderer_count ({renderer_count}) < draw_count ({draw_count}) for config {config_key} ({count} submissions)")
                            
                            # Treat as normal: no flush renderers, all renderers are non-flush
                            flush_count = 0
                            flush_renderers = []
                            non_flush_renderers = renderers
                        else:
                            # Normal case: rdr >= draw
                            flush_count = renderer_count - draw_count
                            flush_renderers = renderers[:flush_count] if flush_count > 0 else []
                            non_flush_renderers = renderers[flush_count:] if flush_count > 0 else renderers
                        
                        # Count non-flush renderers
                        non_flush_counts = {}
                        for r in non_flush_renderers:
                            non_flush_counts[r] = non_flush_counts.get(r, 0) + 1
                        
                        # Count flush renderers
                        flush_counts = {}
                        for r in flush_renderers:
                            flush_counts[r] = flush_counts.get(r, 0) + 1
                        
                        # Build the summary string
                        summary_parts_inner = []
                        
                        # Format non-flush renderers if they exist
                        if non_flush_counts:
                            sorted_non_flush = sorted(non_flush_counts.items(), key=lambda x: x[1], reverse=True)
                            r_summary = ','.join([f"{name}:{cnt}" for name, cnt in sorted_non_flush])
                            summary_parts_inner.append(r_summary)
                        
                        # Add flush renderers with "f:" prefix
                        if flush_counts:
                            sorted_flush = sorted(flush_counts.items(), key=lambda x: x[1], reverse=True)
                            f_summary = ','.join([f"f:{name}:{cnt}" for name, cnt in sorted_flush])
                            summary_parts_inner.append(f_summary)
                        
                        # Special case: no renderers at all
                        if not summary_parts_inner:
                            summary = f"{renderer_count}rdr"
                        else:
                            summary = '|'.join(summary_parts_inner)
                        
                        # Append draw_count if it's an error case (rdr < draw)
                        if renderer_count < draw_count:
                            summary = f"{summary}|draw:{draw_count}"
                        
                        # All configs use sub#id format
                        summary_parts.append(f"sub{idx+1}: {count}[{summary}]")
                    
                    # Join with ", " and add a newline at the end
                    draw_types_map[bench] = ',\n'.join(summary_parts)
                else:
                    draw_types_map[bench] = "No submissions"
                    
            except Exception as e:
                draw_types_map[bench] = f"Error: {str(e)[:30]}"
        else:
            draw_types_map[bench] = f"No JSON"
    
    return draw_types_map

def analyze_ftrace_files(folder_path, benches, folder_exists, folder_name):
    """Main dispatcher for ftrace analysis based on backend type.
    
    Args:
        folder_path: Path to folder containing JSON files
        benches: List of benchmark names
        folder_exists: Boolean indicating if folder exists
        folder_name: Name for display purposes
    
    Returns:
        Dictionary mapping bench to draw_types summary, or None if analysis skipped
    """
    if not folder_exists:
        print(f"\n⚠️  Skipping trace file analysis for '{folder_name}': Folder does not exist")
        return None
    
    backend_type = detect_backend_type(folder_path)
    print(f"  Backend type for '{folder_name}': {backend_type.upper()}")
    
    if backend_type == 'graphite':
        return analyze_ftrace_files_graphite(folder_path, benches)
    else:  # ganesh
        return analyze_ftrace_files_ganesh(folder_path, benches, folder_name)

def generate_ratio_configs(dataframes):
    """Generate ratio configurations dynamically based on available backends.
    
    Creates:
    1. Pairs of any backend vs glesdmsaa (ratio and diff)
    2. Pairs of any backend starting with 'gr' vs grdawn_vk (ratio and diff)
    
    Returns:
        List of ratio configurations
    """
    # Get all backend names (stem of CSV files)
    backends = [Path(csv_file).stem for csv_file in dataframes.keys()]
    
    # Find glesdmsaa backend
    glesdmsaa_backend = None
    for backend in backends:
        if 'glesdmsaa' in backend.lower() or 'gles-dmsaa' in backend.lower():
            glesdmsaa_backend = backend
            break
    
    # Find grdawn_vk backend
    grdawn_vk_backend = None
    for backend in backends:
        if 'grdawn_vk' in backend.lower() or 'grdawn-vk' in backend.lower() or 'grdawnvk' in backend.lower():
            grdawn_vk_backend = backend
            break
    
    ratio_configs = []
    
    # Generate pairs vs glesdmsaa (for all backends except glesdmsaa itself)
    if glesdmsaa_backend:
        for backend in backends:
            if backend != glesdmsaa_backend:
                # Ratio: backend / glesdmsaa
                ratio_configs.append({
                    'name': f'{backend} vs {glesdmsaa_backend} (ratio)',
                    'numerator_patterns': [backend.lower()],
                    'denominator_patterns': [glesdmsaa_backend.lower()],
                    'type': 'ratio'
                })
                # Diff: backend - glesdmsaa
                ratio_configs.append({
                    'name': f'{backend} vs {glesdmsaa_backend} (diff)',
                    'numerator_patterns': [backend.lower()],
                    'denominator_patterns': [glesdmsaa_backend.lower()],
                    'type': 'diff'
                })
    else:
        print("\n⚠️  Warning: 'glesdmsaa' backend not found - skipping vs glesdmsaa comparisons")
    
    # Generate pairs vs grdawn_vk (for all backends starting with 'gr' except grdawn_vk itself)
    if grdawn_vk_backend:
        for backend in backends:
            if backend != grdawn_vk_backend and backend.lower().startswith('gr'):
                # Ratio: backend / grdawn_vk
                ratio_configs.append({
                    'name': f'{backend} vs {grdawn_vk_backend} (ratio)',
                    'numerator_patterns': [backend.lower()],
                    'denominator_patterns': [grdawn_vk_backend.lower()],
                    'type': 'ratio'
                })
                # Diff: backend - grdawn_vk
                ratio_configs.append({
                    'name': f'{backend} vs {grdawn_vk_backend} (diff)',
                    'numerator_patterns': [backend.lower()],
                    'denominator_patterns': [grdawn_vk_backend.lower()],
                    'type': 'diff'
                })
    else:
        print("\n⚠️  Warning: 'grdawn_vk' backend not found - skipping vs grdawn_vk comparisons")
    
    return ratio_configs, glesdmsaa_backend, grdawn_vk_backend

def validate_and_filter_backends(dataframes, ratio_configs):
    """Validate that required backend columns exist for ratio calculations.
    Reports missing columns and filters out invalid ratio configs.
    
    Returns:
        tuple: (filtered_ratio_configs, missing_backends_report)
    """
    available_backends = [Path(csv_file).stem.lower() for csv_file in dataframes.keys()]
    print("\n🔍 Checking backend availability for ratio calculations:")
    print(f"   Available backends: {', '.join(available_backends)}")
    
    missing_report = []
    valid_configs = []
    
    for config in ratio_configs:
        ratio_name = config['name']
        numerator_found = False
        denominator_found = False
        found_numerator = None
        found_denominator = None
        
        # Check numerator
        for pattern in config['numerator_patterns']:
            for backend in available_backends:
                if pattern in backend:
                    numerator_found = True
                    found_numerator = backend
                    break
            if numerator_found:
                break
        
        # Check denominator
        for pattern in config['denominator_patterns']:
            for backend in available_backends:
                if pattern in backend:
                    denominator_found = True
                    found_denominator = backend
                    break
            if denominator_found:
                break
        
        if numerator_found and denominator_found:
            valid_configs.append(config)
            calc_type = config.get('type', 'ratio')
            print(f"   ✅ {ratio_name}: {found_numerator} / {found_denominator} ({calc_type})")
        else:
            missing_parts = []
            if not numerator_found:
                missing_parts.append(f"numerator (patterns: {config['numerator_patterns']})")
            if not denominator_found:
                missing_parts.append(f"denominator (patterns: {config['denominator_patterns']})")
            
            missing_msg = f"   ❌ {ratio_name} - missing {', '.join(missing_parts)}"
            print(missing_msg)
            missing_report.append(f"'{ratio_name}' skipped: missing {', '.join(missing_parts)}")
    
    if missing_report:
        print("\n⚠️  Some ratio columns will not be added due to missing backends")
        for report in missing_report:
            print(f"   {report}")
    
    return valid_configs, missing_report

def find_common_benches(dataframes):
    """Find benches that exist in ALL CSV files and report missing ones."""
    # Get sets of benches from each CSV
    bench_sets = []
    for csv_file, df in dataframes.items():
        bench_set = set(df['bench'].tolist())
        bench_sets.append(bench_set)
        print(f"\n  {Path(csv_file).name}: {len(bench_set)} benchmarks")
    
    # Find intersection (benches in all files)
    common_benches = set.intersection(*bench_sets) if bench_sets else set()
    
    # Report missing benchmarks
    print("\n📊 Benchmark Analysis:")
    print(f"  Total unique benchmarks across all files: {len(set.union(*bench_sets))}")
    print(f"  Benchmarks present in ALL files: {len(common_benches)}")
    
    # Check each file for missing benchmarks
    all_benches_union = set.union(*bench_sets)
    
    # Track duplicates across files (benchmark names that appear multiple times in same file)
    duplicate_across_files = {}
    for csv_file, df in dataframes.items():
        dup_in_file = df[df['bench'].duplicated(keep=False)]['bench'].tolist()
        if dup_in_file:
            duplicate_across_files[Path(csv_file).name] = set(dup_in_file)
    
    if duplicate_across_files:
        print("\n⚠️  DUPLICATE BENCHMARKS DETECTED ACROSS FILES:")
        for file_name, benches in duplicate_across_files.items():
            print(f"  • {file_name}: {', '.join(benches)} (appears multiple times - handled by keeping first)")
    
    # Report missing benchmarks per file
    missing_benchmarks_report = []
    for bench in sorted(all_benches_union):
        if bench not in common_benches:
            missing_in = []
            for csv_file, bench_set in zip(dataframes.keys(), bench_sets):
                if bench not in bench_set:
                    missing_in.append(Path(csv_file).name)
            missing_msg = f"  ⚠️  '{bench}' - missing in: {', '.join(missing_in)}"
            print(missing_msg)
            missing_benchmarks_report.append(missing_msg)
    
    return common_benches, missing_benchmarks_report

def create_comparison_page(dataframes, draw_types_maps, folder_paths, common_benches):
    """Create the comparison dataframe with summary columns for each folder.
    Only includes benchmarks that exist in ALL CSV files.
    """
    # Convert to sorted list for consistent ordering
    benches = sorted(list(common_benches))
    
    if not benches:
        print("\n❌ Error: No common benchmarks found across all CSV files!")
        print("   Please ensure at least one benchmark name appears in all CSV files.")
        sys.exit(1)
    
    # Create ordered ID column
    ordered_ids = list(range(1, len(benches) + 1))
    
    # Prepare comparison data
    comparison_data = {
        'ID': ordered_ids,
        'Bench': benches
    }
    
    # Add mean columns for each CSV (named after filename/backend)
    for csv_file, df in dataframes.items():
        backend_name = Path(csv_file).stem  # Remove .csv extension
        # Create dictionary only for common benches (first occurrence if duplicates existed)
        mean_dict = dict(zip(df['bench'], df['mean']))
        comparison_data[backend_name] = [mean_dict.get(bench, float('nan')) for bench in benches]
    
    # Add summary columns for each folder
    for idx, folder_path in enumerate(folder_paths):
        folder_name = Path(folder_path).name
        if folder_name == '.' or folder_name == '/':
            folder_name = f"folder_{idx + 1}"
        
        # Get the draw_types map for this folder (None if analysis was skipped)
        draw_types_map = draw_types_maps[idx] if idx < len(draw_types_maps) else None
        
        # Column name: "summary of [folder_name]"
        summary_col_name = f"summary of {folder_name}"
        
        if draw_types_map is not None:
            comparison_data[summary_col_name] = [draw_types_map.get(bench, "Bench not found in trace files") for bench in benches]
        else:
            comparison_data[summary_col_name] = ["Trace analysis skipped (folder not found)" for _ in benches]
    
    # Create dataframe
    comparison_df = pd.DataFrame(comparison_data)
    
    return comparison_df

def add_excel_ratio_formulas(writer, comparison_df, dataframes, valid_ratio_configs):
    """Add Excel formulas for valid ratio columns only."""
    workbook = writer.book
    sheet = workbook['comparison']
    
    # Get all headers
    headers = [cell.value for cell in sheet[1]]
    
    # Find column indices for each backend
    column_map = {}
    for col_idx, header in enumerate(headers, 1):
        if header and isinstance(header, str):
            header_lower = header.lower()
            column_map[header_lower] = col_idx
    
    # Process each valid ratio configuration
    added_ratios = []
    for config in valid_ratio_configs:
        ratio_name = config['name']
        calc_type = config.get('type', 'ratio')
        
        # Find numerator column
        numerator_col = None
        numerator_name = None
        for pattern in config['numerator_patterns']:
            for col_name, col_idx in column_map.items():
                if pattern in col_name:
                    numerator_col = col_idx
                    numerator_name = [h for h in headers if column_map.get(h.lower()) == col_idx][0]
                    break
            if numerator_col:
                break
        
        # Find denominator column
        denominator_col = None
        denominator_name = None
        for pattern in config['denominator_patterns']:
            for col_name, col_idx in column_map.items():
                if pattern in col_name:
                    denominator_col = col_idx
                    denominator_name = [h for h in headers if column_map.get(h.lower()) == col_idx][0]
                    break
            if denominator_col:
                break
        
        if not numerator_col or not denominator_col:
            print(f"\n⚠️  Warning: Could not find columns for '{ratio_name}' (should have been filtered)")
            continue
        
        # Check if ratio column already exists
        if ratio_name in headers:
            ratio_col_idx = headers.index(ratio_name) + 1
        else:
            # Add column before summary columns if possible
            summary_cols = [col for col in headers if col and col.startswith('summary of')]
            if summary_cols:
                first_summary_idx = headers.index(summary_cols[0]) + 1
                sheet.insert_cols(first_summary_idx)
                ratio_col_idx = first_summary_idx
                sheet.cell(row=1, column=ratio_col_idx, value=ratio_name)
                # Update headers list
                headers.insert(first_summary_idx - 1, ratio_name)
            else:
                # Add at the end
                ratio_col_idx = len(headers) + 1
                sheet.cell(row=1, column=ratio_col_idx, value=ratio_name)
        
        # Add Excel formula for each row
        print(f"\nAdding Excel formula for column: {ratio_name}")
        if calc_type == 'diff':
            print(f"  = {numerator_name} - {denominator_name}")
        else:
            print(f"  = {numerator_name} / {denominator_name}")
        
        formula_count = 0
        for row_idx in range(2, len(comparison_df) + 2):
            # Create Excel formula
            if calc_type == 'diff':
                formula = f"={get_column_letter(numerator_col)}{row_idx}-{get_column_letter(denominator_col)}{row_idx}"
                number_format = "0.000"
            else:  # ratio
                formula = f"={get_column_letter(numerator_col)}{row_idx}/{get_column_letter(denominator_col)}{row_idx}"
                number_format = "0.000"
            
            # Set the formula in the cell
            cell = sheet.cell(row=row_idx, column=ratio_col_idx)
            cell.value = formula
            cell.number_format = number_format
            
            # Add a comment explaining the formula
            if calc_type == 'diff':
                cell.comment = Comment(f"Formula: {numerator_name} - {denominator_name}", "Script")
            else:
                cell.comment = Comment(f"Formula: {numerator_name} / {denominator_name}", "Script")
            formula_count += 1
        
        # Auto-adjust column width
        column_letter = get_column_letter(ratio_col_idx)
        sheet.column_dimensions[column_letter].width = 20
        
        print(f"  - Added formula to {formula_count} rows")
        print(f"  - Column: {column_letter}")
        print(f"  - Example formula: {formula}")
        added_ratios.append(ratio_name)
    
    return added_ratios

def apply_table_formatting(writer, comparison_df):
    """Apply Excel table formatting to the comparison sheet for sorting/filtering."""
    workbook = writer.book
    sheet = workbook['comparison']
    
    # Determine the range of the table
    start_row = 1
    start_col = 1
    end_row = len(comparison_df) + 1  # +1 for header
    
    # Find the last column with data
    max_col = 1
    for row in sheet.iter_rows(max_row=1):
        for cell in row:
            if cell.value:
                max_col = max(max_col, cell.column)
    
    # Create table range reference
    table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(max_col)}{end_row}"
    
    # Create table
    table_name = "ComparisonTable"
    
    # Remove existing table if it exists
    if table_name in sheet.tables:
        del sheet.tables[table_name]
    
    table = Table(displayName=table_name, ref=table_range)
    
    # Define table style
    style = TableStyleInfo(
        name="TableStyleMedium9",  # Clean, professional style
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,  # Alternating row colors
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # Add table to sheet
    sheet.add_table(table)
    
    print(f"\n📊 Applied Excel table formatting to range: {table_range}")
    print(f"   - Table name: {table_name}")
    print("   - Features: Sorting, Filtering, Alternating row colors")
    
    return table

def format_excel_workbook(writer, comparison_df, dataframes, folder_paths, missing_backends_report):
    """Format the Excel workbook with styling."""
    workbook = writer.book
    
    # Add missing backends information to comparison sheet
    if missing_backends_report:
        comparison_sheet = workbook['comparison']
        # Add a comment to cell A1 about missing backends
        missing_text = "⚠️ Missing Backends for Ratio/Diff Columns:\n\n" + "\n".join(missing_backends_report)
        comment = Comment(missing_text, "Data Processor")
        comparison_sheet['A1'].comment = comment
    
    # Format comparison sheet
    comparison_sheet = workbook['comparison']
    
    # Apply table formatting (this enables sorting/filtering)
    apply_table_formatting(writer, comparison_df)
    
    # Auto-adjust column widths (do this after table creation)
    for column in comparison_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        comparison_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row and first column for easy scrolling
    comparison_sheet.freeze_panes = comparison_sheet['B2']  # Freeze row 1 and column A
    
    print(f"   - Frozen panes: Row 1 (headers) and Column A (ID/Bench)")
    
    # Format individual CSV sheets
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for csv_file in dataframes.keys():
        sheet_name = Path(csv_file).stem[:31]  # Excel sheet name max 31 chars
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Style header
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Freeze header row for CSV sheets too
            sheet.freeze_panes = sheet['A2']

def print_summary(folder_paths, folder_exists_list, draw_types_maps, dataframes, comparison_df, 
                 added_ratios, missing_benchmarks_report, missing_backends_report,
                 glesdmsaa_backend, grdawn_vk_backend):
    """Print a summary of the analysis."""
    print("\n" + "="*60)
    print("✅ Analysis complete!")
    
        
    if folder_paths:
        print("\n📁 Folders analyzed:")
        for idx, folder_path in enumerate(folder_paths):
            folder_name = Path(folder_path).name
            exists = folder_exists_list[idx] if idx < len(folder_exists_list) else False
            status = "✅ Analyzed" if exists else "⚠️  Skipped (not found)"
            print(f"  {idx + 1}. {folder_name}: {status}")
            
            # Show statistics for this folder if analysis was done
            if exists and draw_types_maps and idx < len(draw_types_maps) and draw_types_maps[idx] is not None:
                benches_with_info = sum(1 for v in draw_types_maps[idx].values() 
                                       if not v.startswith("JSON file not found") 
                                       and not v.startswith("Error")
                                       and v != "Bench not found in trace files")
                total_benches = len(draw_types_maps[idx])
                print(f"     Benchmarks with trace info: {benches_with_info}/{total_benches}")
    else:
        print("\n📁 No folders provided - trace analysis skipped")
    
    print(f"\n📊 CSV files processed: {len(dataframes)}")
    for csv_file, df in dataframes.items():
        print(f"  - {Path(csv_file).name}: {len(df)} rows")
    
    print(f"\n📈 Comparison page: {len(comparison_df)} benchmarks (common across all files)")
    print(f"   Base columns: ID, Bench, backend columns ({len(dataframes)}), summary columns ({len(folder_paths)})")
    
    # Show backend reference
    if glesdmsaa_backend:
        print(f"   Reference backend (glesdmsaa): {glesdmsaa_backend}")
    if grdawn_vk_backend:
        print(f"   Reference backend (grdawn_vk): {grdawn_vk_backend}")
    
    ratio_count = len([c for c in added_ratios if '(ratio)' in c])
    diff_count = len([c for c in added_ratios if '(diff)' in c])
    print(f"   Comparison columns added: {len(added_ratios)} total ({ratio_count} ratios, {diff_count} diffs)")
    
    # Report missing benchmarks
    if missing_benchmarks_report:
        print(f"\n⚠️  MISSING BENCHMARKS ({len(missing_benchmarks_report)}):")
        for report in missing_benchmarks_report[:10]:  # Show first 10
            print(f"  {report}")
        if len(missing_benchmarks_report) > 10:
            print(f"  ... and {len(missing_benchmarks_report) - 10} more")
    
    # Report missing backends
    if missing_backends_report:
        print(f"\n⚠️  MISSING BACKENDS FOR COMPARISONS ({len(missing_backends_report)}):")
        for report in missing_backends_report:
            print(f"  {report}")
    
    # Note about JSON file lookup
    print("\n📁 JSON File Lookup:")
    print("  - Dots (.) are automatically removed from benchmark names when locating JSON files")
    print("  - The script tries sanitized name first, then original name")
    print("  - Example: 'bench.test.1' -> looks for 'benchestest1.json', then 'bench.test.1.json'")
    
    # Note about table features
    print("\n📊 Excel Table Features Available:")
    print("  - Click drop-down arrows in headers to sort/filter")
    print("  - Alternating row colors for easy reading")
    print("  - Header row frozen for scrolling")
    print("  - Resizable and customizable table style")
    
    # Note about duplicate handling
    print("\n⚠️  Duplicate Benchmark Handling:")
    print("  - Duplicate benchmark names in CSV files are detected and reported in console output")
    print("  - Only the first occurrence of each duplicate is used in comparisons")
    print("  - No error messages are written to Excel sheets")
    
    # Note about comparison columns
    if added_ratios:
        print("\n📐 Comparison Columns Added (Excel Formulas):")
        ratio_cols = [c for c in added_ratios if '(ratio)' in c]
        diff_cols = [c for c in added_ratios if '(diff)' in c]
        if ratio_cols:
            print("   Ratios (division):")
            for col in ratio_cols[:5]:  # Show first 5
                print(f"    - {col}")
            if len(ratio_cols) > 5:
                print(f"    ... and {len(ratio_cols) - 5} more")
        if diff_cols:
            print("   Diffs (subtraction):")
            for col in diff_cols[:5]:  # Show first 5
                print(f"    - {col}")
            if len(diff_cols) > 5:
                print(f"    ... and {len(diff_cols) - 5} more")
    else:
        print("\n⚠️  No comparison columns were added due to missing backend requirements")

def main():
    """Main function to orchestrate the script."""
    print("Starting benchmark analysis...")
    print("="*60)
    
    # Validate arguments and get folder paths and CSV files
    folder_paths, csv_files, folder_exists_list = validate_arguments()
    
    print(f"\n📁 Folder paths ({len(folder_paths)}):")
    for folder_path in folder_paths:
        print(f"  - {folder_path}")
    
    print(f"\n📄 CSV files ({len(csv_files)}):")
    for csv_file in csv_files:
        print(f"  - {csv_file}")
    
    # Read CSV files (with duplicate detection)
    print("\n📖 Reading CSV files and checking for duplicates...")
    dataframes = read_csv_files(csv_files)
    
    # Generate ratio configurations dynamically
    print("\n🔧 Generating comparison configurations...")
    ratio_configs, glesdmsaa_backend, grdawn_vk_backend = generate_ratio_configs(dataframes)
    
    if ratio_configs:
        print(f"\n   Generated {len(ratio_configs)} comparison configurations:")
        for config in ratio_configs:
            calc_type = config.get('type', 'ratio')
            print(f"    - {config['name']} ({calc_type})")
    else:
        print("\n   ⚠️  No comparison configurations generated!")
        print("   Please ensure you have at least glesdmsaa backend or grdawn_vk backend available")
    
    # Validate backends and filter ratio configs
    valid_ratio_configs, missing_backends_report = validate_and_filter_backends(dataframes, ratio_configs)
    
    # Find common benches across all CSV files and report missing ones
    print("\n🔍 Finding common benchmarks across all CSV files...")
    common_benches, missing_benchmarks_report = find_common_benches(dataframes)
    
    # Analyze ftrace JSON files for each folder (only for common benches)
    print("\n🔍 Analyzing ftrace JSON files...")
    draw_types_maps = []
    for idx, folder_path in enumerate(folder_paths):
        folder_name = Path(folder_path).name
        print(f"\n  Processing folder {idx + 1}: {folder_name}")
        folder_exists = folder_exists_list[idx] if idx < len(folder_exists_list) else False
        draw_types_map = analyze_ftrace_files(folder_path, common_benches, folder_exists, folder_name)
        draw_types_maps.append(draw_types_map)
    
    # Create comparison dataframe with summary columns (only common benches)
    print("\n📊 Creating comparison page...")
    comparison_df = create_comparison_page(dataframes, draw_types_maps, folder_paths, common_benches)
    
    # Generate output filename
    output_file = "backend_comparison.xlsx"
    
    # Write to Excel with multiple sheets
    print(f"\n💾 Generating Excel workbook: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write comparison sheet as first sheet
        comparison_df.to_excel(writer, sheet_name='comparison', index=False)
        
        # Write individual CSV sheets (original full data)
        for csv_file, df in dataframes.items():
            sheet_name = Path(csv_file).stem[:31]  # Excel sheet name max 31 chars
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Add Excel formulas for valid ratio columns only
        added_ratios = add_excel_ratio_formulas(writer, comparison_df, dataframes, valid_ratio_configs)
        
        # Apply formatting (including table formatting)
        format_excel_workbook(writer, comparison_df, dataframes, folder_paths, missing_backends_report)
    
    # Print summary
    print_summary(folder_paths, folder_exists_list, draw_types_maps, dataframes, comparison_df, 
                 added_ratios, missing_benchmarks_report, missing_backends_report,
                 glesdmsaa_backend, grdawn_vk_backend)
    
    # Show sample of the comparison table structure (only for non-formula columns)
    print("\n📋 Sample of comparison table structure (first 5 rows, values from CSV data only):")
    display_cols = ['ID', 'Bench']
    
    # Add first two backend columns for sample (these contain actual values)
    backend_cols = [col for col in comparison_df.columns 
                   if col not in ['ID', 'Bench'] 
                   and not col.startswith('summary of')][:2]
    display_cols.extend(backend_cols)
    
    # Add first summary column (contains trace analysis results)
    summary_cols = [col for col in comparison_df.columns if col.startswith('summary of')]
    if summary_cols:
        display_cols.append(summary_cols[0])
    
    # Display the dataframe sample (without ratio columns which are Excel-only formulas)
    if len(comparison_df) > 0:
        print(comparison_df[display_cols].head().to_string())
        print("\n   Note: Comparison columns (ratios and diffs) are Excel formulas and not shown in this preview.")
        print("         They will appear when you open the Excel file and will automatically calculate.")
    else:
        print("\n   ⚠️  No common benchmarks found to display!")
    
    print("\n📁 Output file details:")
    print(f"  - File: {output_file}")
    print(f"  - Total sheets: {len(dataframes) + 1}")
    print(f"  - First sheet: comparison ({len(comparison_df)} benchmarks)")
    print(f"  - Summary columns: {len(summary_cols)}")
    for col in summary_cols:
        print(f"    * {col}")
    print(f"  - Comparison columns: {len(added_ratios)} (Excel formulas)")
    ratio_cols = [c for c in added_ratios if '(ratio)' in c]
    diff_cols = [c for c in added_ratios if '(diff)' in c]
    if ratio_cols:
        print(f"    * Ratios: {len(ratio_cols)} columns")
    if diff_cols:
        print(f"    * Diffs: {len(diff_cols)} columns")
    
    print("\n💡 Tips for using the Excel file:")
    print("  1. Use the drop-down arrows in any header to sort or filter data")
    print("  2. Comparison columns include both ratios (division) and diffs (subtraction)")
    print("  3. The first row and column are frozen for easy scrolling")
    print("  4. Table formatting updates automatically when you add/remove data")
    print("  5. All formulas recalculate when source data changes")
    print("  6. Summary columns show trace analysis results (draw types per flush)")
    print("  7. Individual CSV sheets contain ALL original data, comparison sheet only shows common benchmarks")
    print("  8. Dots in benchmark names are removed when locating JSON files (e.g., 'test.bench' -> 'testbench.json')")
    print("  9. Duplicate benchmark warnings are shown in console output only, not in Excel sheets")
    print("  10. Reference backends: glesdmsaa and grdawn_vk are used as baselines for comparisons")
    
    print("\n" + "="*60)

if __name__ == "__main__":
    main()