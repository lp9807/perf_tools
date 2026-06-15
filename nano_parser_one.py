#!/usr/bin/env python3
"""
================================================================================
Benchmark Analysis Tool - Version 2.7
================================================================================
Script to process multiple Excel files and generate comparison with version tracking.

Version History:
- v1.0: Initial release with CSV file inputs
- v2.0: Major redesign - switched to Excel file input, optional folders, 3 ratio columns
- v2.1: Multiple Excel files support, API version detection, cross-version comparisons
- v2.2: Reorganized column order - backend columns first, then ratio columns
- v2.3: Simplified column names when only one input file (no version suffix)
- v2.4: Backup original backend pages into output workbook as separate sheets
- v2.5: Separate comparison pages per version + dedicated cross-version page
- v2.6: Extract summary columns from existing comparison page and append to new pages
- v2.7: Report issues AND filter benchmarks to only include those existing in all backends

Features:
- Accepts 0-2 folder paths for trace analysis (optional)
- Accepts MULTIPLE Excel files with multiple sheets (one per backend)
- Detects API version from filename (api[0-9]+ pattern)
- Each sheet must contain 'bench' and 'mean' columns
- Generates separate comparison page for EACH version
- Creates dedicated cross-version comparison page
- Adds version-specific ratio columns (3 ratios per version)
- Optional ftrace JSON analysis for draw types
- Auto-increments output filename version number
- Backs up original backend pages
- Extracts summary columns from existing comparison page and appends to new pages
- Reports benchmark cases that don't exist in all pages
- Detects and reports duplicate benchmark names within each sheet
- Reports columns (backends) that don't exist in all versions
- **NEW: Filters benchmarks to only include those that exist in ALL backends per version**

Usage: python script.py [<folder_path1>] [<folder_path2>] <excel_file1.xlsx> [<excel_file2.xlsx> ...]

Author: Benchmark Analysis Tool
Version: 2.7
Date: 2026-06-12
================================================================================
"""

import sys
import os
import json
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# Version constant - Minor version increment to 2.7
MAJOR_VERSION = 2
MINOR_VERSION = 7
VERSION = f"{MAJOR_VERSION}.{MINOR_VERSION}"

def get_next_version_number(base_filename):
    """Get the next available version number for output filename."""
    pattern = re.compile(rf'{re.escape(base_filename)}_v(\d+)\.xlsx$')
    existing_versions = []
    
    # Check for existing files with version numbers
    for file in Path('.').glob(f'{base_filename}_v*.xlsx'):
        match = pattern.match(file.name)
        if match:
            existing_versions.append(int(match.group(1)))
    
    if existing_versions:
        next_version = max(existing_versions) + 1
    else:
        next_version = 1
    
    return next_version

def generate_output_filename(excel_files, folder_paths):
    """Generate output filename with version number."""
    # Determine base name based on input
    if len(excel_files) == 1:
        # Single file: use basename + "_benchmark_comparison"
        base_name = f"{Path(excel_files[0]).stem}_benchmark_comparison"
    else:
        # Multiple files: use benchmark_comparison
        base_name = "benchmark_comparison"
    
    # Get next version number
    version_num = get_next_version_number(base_name)
    
    # Create filename
    output_file = f"{base_name}_v{version_num}.xlsx"
    
    return output_file, version_num

def validate_arguments():
    """Validate command line arguments."""
    if len(sys.argv) < 2:
        print("Error: At least one argument required (Excel file)")
        print(f"Usage: {sys.argv[0]} [<folder_path1>] [<folder_path2>] <excel_file1.xlsx> [<excel_file2.xlsx> ...]")
        print("Note: Folder parameters are optional (0-2 folders), followed by at least one Excel file")
        sys.exit(1)
    
    # Parse arguments: first 0-2 could be folders, rest are Excel files
    folder_paths = []
    excel_files = []
    
    # Check each argument
    for arg in sys.argv[1:]:
        if arg.lower().endswith('.xlsx') or arg.lower().endswith('.xls'):
            excel_files.append(arg)
        elif len(folder_paths) < 2:  # Max 2 folders
            folder_paths.append(arg)
        else:
            print(f"Error: Maximum 2 folder parameters allowed. Extra parameter: '{arg}'")
            sys.exit(1)
    
    if len(excel_files) == 0:
        print("Error: At least one Excel file (.xlsx or .xls) must be provided")
        print(f"Usage: {sys.argv[0]} [<folder_path1>] [<folder_path2>] <excel_file1.xlsx> [<excel_file2.xlsx> ...]")
        sys.exit(1)
    
    # Check if folders exist (warn but don't exit)
    folder_exists = []
    for folder_path in folder_paths:
        exists = os.path.isdir(folder_path)
        folder_exists.append(exists)
        if not exists:
            print(f"Warning: Folder path '{folder_path}' does not exist - trace analysis will be skipped for this folder")
    
    # Check Excel files exist
    for excel_file in excel_files:
        if not os.path.isfile(excel_file):
            print(f"Error: Excel file '{excel_file}' does not exist")
            sys.exit(1)
    
    return folder_paths, excel_files, folder_exists

def extract_api_version(filename):
    """Extract API version from filename using pattern api[0-9]+."""
    match = re.search(r'api(\d+)', filename, re.IGNORECASE)
    if match:
        return f"api{match.group(1)}"
    return None

def detect_backend_type(folder_path):
    """Detect backend type from folder name."""
    folder_name = Path(folder_path).name.lower()
    if 'gr' in folder_name:
        return 'graphite'
    else:
        return 'ganesh'

def extract_summary_columns_from_comparison(excel_file):
    """Extract summary columns from existing comparison sheet if it exists."""
    try:
        # Load the workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Check if comparison sheet exists
        if 'comparison' not in wb.sheetnames:
            print(f"  No 'comparison' sheet found in {Path(excel_file).name}")
            return None
        
        # Read the comparison sheet
        df = pd.read_excel(excel_file, sheet_name='comparison')
        
        # Find columns that are summary columns (contain 'summary' in name or are not standard columns)
        standard_columns = ['ID', 'Bench']
        summary_columns = {}
        
        for col in df.columns:
            if col not in standard_columns:
                # Check if it looks like a summary column
                if 'summary' in col.lower() or 'draw' in col.lower() or 'trace' in col.lower():
                    summary_columns[col] = df[col].tolist()
                    print(f"  - Found summary column: '{col}'")
        
        if summary_columns:
            print(f"  Extracted {len(summary_columns)} summary columns from existing comparison page")
            return summary_columns
        else:
            print(f"  No summary columns found in existing comparison page")
            return None
            
    except Exception as e:
        print(f"  Error reading comparison sheet from {Path(excel_file).name}: {e}")
        return None

def check_columns_coverage(version_groups, all_backends):
    """Check and report columns (backends) that don't exist in all versions."""
    print("\n" + "="*60)
    print("📊 Column/Backend Coverage Analysis")
    print("="*60)
    
    # For each backend, track which versions have it
    backend_coverage = defaultdict(list)
    
    for version_name, version_data in version_groups.items():
        # Get all backend names (without version suffix) for this version
        version_backends = set()
        for col_name in version_data['dataframes'].keys():
            # Remove version suffix to get backend name
            backend_name = col_name.rsplit('_', 1)[0] if '_' in col_name else col_name
            version_backends.add(backend_name)
        
        for backend in version_backends:
            backend_coverage[backend].append(version_name)
    
    # Find backends missing from some versions
    all_versions = set(version_groups.keys())
    missing_backends = {}
    
    for backend, present_versions in backend_coverage.items():
        present_set = set(present_versions)
        missing = all_versions - present_set
        if missing:
            missing_backends[backend] = {
                'present_in': sorted(present_versions),
                'missing_in': sorted(missing)
            }
    
    if missing_backends:
        print("\n⚠️  Backends missing from some versions:")
        for backend, info in sorted(missing_backends.items()):
            print(f"\n  Backend '{backend}':")
            print(f"    - Present in: {', '.join(info['present_in'])}")
            print(f"    - Missing in: {', '.join(info['missing_in'])}")
    else:
        print("\n✅ All backends are present in all versions!")
    
    # Also find backends that are only in a single version
    single_version_backends = {b: v for b, v in backend_coverage.items() if len(v) == 1}
    if single_version_backends:
        print(f"\n📌 Backends unique to a single version ({len(single_version_backends)}):")
        for backend, versions in sorted(single_version_backends.items()):
            print(f"    - '{backend}' (only in {versions[0]})")
    
    print("="*60)
    
    return missing_backends, single_version_backends

def check_duplicate_benchmarks(version_groups):
    """Check and report duplicate benchmark names within each sheet."""
    print("\n" + "="*60)
    print("🔍 Duplicate Benchmark Detection")
    print("="*60)
    
    duplicates_found = False
    duplicate_report = {}
    
    for version_name, version_data in version_groups.items():
        print(f"\n  Checking version: {version_name}")
        version_duplicates = {}
        
        for sheet_name, df in version_data['sheets'].items():
            # Check for duplicates in the bench column
            bench_counts = df['bench'].value_counts()
            duplicates = bench_counts[bench_counts > 1]
            
            if len(duplicates) > 0:
                duplicates_found = True
                version_duplicates[sheet_name] = duplicates.to_dict()
                print(f"    ⚠️  Sheet '{sheet_name}' has {len(duplicates)} duplicate benchmark(s):")
                for bench_name, count in duplicates.items():
                    print(f"        - '{bench_name}' appears {count} times")
            else:
                print(f"    ✓ Sheet '{sheet_name}' has no duplicate benchmarks")
        
        if version_duplicates:
            duplicate_report[version_name] = version_duplicates
    
    if not duplicates_found:
        print("\n✅ No duplicate benchmarks found in any sheet!")
    
    print("="*60)
    
    return duplicate_report

def check_missing_benchmarks(version_groups):
    """Check and report benchmark cases that don't exist in all pages."""
    print("\n" + "="*60)
    print("📋 Benchmark Coverage Analysis")
    print("="*60)
    
    # Get all unique benchmarks across all versions
    all_benchmarks = set()
    for version_data in version_groups.values():
        for df in version_data['dataframes'].values():
            all_benchmarks.update(df['bench'].tolist())
    
    # For each version, track which benchmarks are present
    version_benchmarks = {}
    for version_name, version_data in version_groups.items():
        benchmarks = set()
        for df in version_data['dataframes'].values():
            benchmarks.update(df['bench'].tolist())
        version_benchmarks[version_name] = benchmarks
    
    # Find benchmarks missing from each version
    missing_report = {}
    for version_name, benchmarks in version_benchmarks.items():
        missing = all_benchmarks - benchmarks
        if missing:
            missing_report[version_name] = missing
    
    if missing_report:
        print("\n⚠️  Missing Benchmarks Detected:")
        for version_name, missing in sorted(missing_report.items()):
            print(f"\n  Version '{version_name}' is missing {len(missing)} benchmark(s):")
            for bench in sorted(missing):
                # Find which versions have this benchmark
                present_in = []
                for v, b in version_benchmarks.items():
                    if bench in b:
                        present_in.append(v)
                if present_in:
                    print(f"    - {bench} (present in: {', '.join(present_in)})")
                else:
                    print(f"    - {bench} (present in: none)")
    else:
        print("\n✅ All benchmarks are present in all versions!")
    
    # Also check for benchmarks that are unique to a single version
    benchmark_versions = defaultdict(list)
    for version_name, benchmarks in version_benchmarks.items():
        for bench in benchmarks:
            benchmark_versions[bench].append(version_name)
    
    unique_benchmarks = {bench: versions for bench, versions in benchmark_versions.items() if len(versions) == 1}
    if unique_benchmarks:
        print(f"\n📌 Benchmarks unique to a single version ({len(unique_benchmarks)}):")
        for bench, versions in sorted(unique_benchmarks.items()):
            print(f"    - {bench} (only in {versions[0]})")
    
    print("="*60)
    
    return missing_report, unique_benchmarks, version_benchmarks

def read_excel_sheets(excel_file, api_version):
    """Read all sheets from Excel file, excluding 'comparison' sheet."""
    try:
        # Dictionary to store dataframes with backend names
        dataframes = {}
        # Dictionary to store original sheet data for backup
        original_sheets = {}
        
        # Load the workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Iterate through all sheets
        for sheet_name in wb.sheetnames:
            # Skip the 'comparison' sheet if it exists
            if sheet_name.lower() == 'comparison':
                print(f"  - Skipping existing 'comparison' sheet in {Path(excel_file).name}")
                continue
            
            # Read sheet into DataFrame
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Validate required columns
            if 'bench' not in df.columns or 'mean' not in df.columns:
                print(f"  Warning: Sheet '{sheet_name}' in {Path(excel_file).name} missing 'bench' or 'mean' column - skipping")
                continue
            
            # Store original sheet data for backup
            original_sheets[sheet_name] = df
            
            # Create column name with version
            column_name = f"{sheet_name}_{api_version}" if api_version else sheet_name
            dataframes[column_name] = df
            print(f"  - Loaded sheet '{sheet_name}' -> column '{column_name}': {len(df)} rows")
        
        return dataframes, original_sheets
        
    except Exception as e:
        print(f"Error reading Excel file '{excel_file}': {e}")
        return None, None

def read_multiple_excel_files(excel_files):
    """Read all Excel files and combine their data."""
    all_dataframes = {}
    all_backends = set()
    version_info = {}
    all_original_sheets = {}
    version_groups = defaultdict(dict)  # Group data by API version
    
    print(f"\n📖 Reading Excel files...")
    
    # Extract summary columns from the first Excel file's comparison page
    summary_columns = None
    if excel_files:
        print(f"\n📋 Checking first Excel file for existing comparison page: {Path(excel_files[0]).name}")
        summary_columns = extract_summary_columns_from_comparison(excel_files[0])
    
    for excel_file in excel_files:
        filename = Path(excel_file).stem
        api_version = extract_api_version(filename)
        
        if api_version:
            print(f"\n📖 Reading Excel file: {Path(excel_file).name} (detected version: {api_version})")
        else:
            print(f"\n📖 Reading Excel file: {Path(excel_file).name} (no version detected)")
            api_version = "default"
        
        dataframes, original_sheets = read_excel_sheets(excel_file, api_version)
        
        if dataframes and original_sheets:
            all_dataframes.update(dataframes)
            
            # Store original sheets with version info
            for sheet_name, df in original_sheets.items():
                backup_sheet_name = f"{sheet_name}_{api_version}"
                all_original_sheets[backup_sheet_name] = df
            
            # Group data by version
            version_groups[api_version]['dataframes'] = dataframes
            version_groups[api_version]['file'] = excel_file
            version_groups[api_version]['sheets'] = original_sheets
            
            version_info[api_version] = {
                'file': excel_file,
                'columns': list(dataframes.keys())
            }
            
            # Collect unique backend names
            for col_name in dataframes.keys():
                backend = col_name.rsplit('_', 1)[0] if '_' in col_name else col_name
                all_backends.add(backend)
    
    if len(all_dataframes) == 0:
        print("Error: No valid data loaded from any Excel file")
        sys.exit(1)
    
    # Check for missing/extra backends across versions
    missing_backends, single_version_backends = check_columns_coverage(version_groups, all_backends)
    
    # Check for duplicate benchmarks
    duplicate_report = check_duplicate_benchmarks(version_groups)
    
    # Check for missing benchmarks
    missing_report, unique_benchmarks, version_benchmarks = check_missing_benchmarks(version_groups)
    
    print(f"\n📊 Summary: Loaded {len(all_dataframes)} backend columns from {len(excel_files)} files")
    print(f"   Backends found: {', '.join(sorted(all_backends))}")
    print(f"   Versions found: {', '.join(version_groups.keys())}")
    print(f"   Original sheets to backup: {len(all_original_sheets)}")
    
    return (all_dataframes, all_backends, version_info, version_groups, 
            all_original_sheets, summary_columns, missing_report, 
            unique_benchmarks, duplicate_report, missing_backends, 
            single_version_backends, version_benchmarks)

def analyze_ftrace_files_ganesh(folder_path, benches, folder_name):
    """Analyze ftrace JSON files for Ganesh backend."""
    folder = Path(folder_path)
    draw_types_map = {}
    
    # Pattern to match SurfaceDrawContext::draw* functions
    draw_pattern = re.compile(r'ScalerContext::draw([a-zA-Z]+)')
    # Pattern to match GrDrawingManager::flush specifically
    flush_pattern = re.compile(r'GrDrawingManager::flush')
    
    for bench in benches:
        json_file = folder / f"{bench}.json"
        draw_counts = {}
        flush_count = 0
        
        if json_file.exists():
            try:
                with open(json_file, 'r') as f:
                    data = json.load(f)
                
                # Handle different ftrace JSON structures
                def process_trace(trace_data):
                    nonlocal flush_count
                    if isinstance(trace_data, dict):
                        # Check for function name in common fields
                        func_name = trace_data.get('func') or trace_data.get('function') or trace_data.get('name')
                        if func_name:
                            # Count draw types
                            draw_match = draw_pattern.search(func_name)
                            if draw_match:
                                draw_type = draw_match.group(1)
                                draw_counts[draw_type] = draw_counts.get(draw_type, 0) + 1
                            
                            # Count GrDrawingManager::flush calls
                            if flush_pattern.search(func_name):
                                flush_count += 1
                        
                        # Recursively process nested structures
                        for value in trace_data.values():
                            if isinstance(value, (dict, list)):
                                process_trace(value)
                    elif isinstance(trace_data, list):
                        for item in trace_data:
                            process_trace(item)
                
                process_trace(data)
                
                if draw_counts and flush_count > 0:
                    # Calculate division result for each draw type: total_count / flush_count
                    sorted_counts = sorted(draw_counts.items(), key=lambda x: x[1], reverse=True)
                    result_parts = []
                    for draw_type, count in sorted_counts:
                        division_result = count / flush_count
                        # Format with 2 decimal places for the division result
                        result_parts.append(f"{draw_type}:{division_result:.2f}")
                    draw_types_map[bench] = ', '.join(result_parts)
                elif draw_counts:
                    draw_types_map[bench] = f"No flush calls found. Draw types: {', '.join([f'{dt}:{cnt}' for dt, cnt in draw_counts.items()])}"
                else:
                    draw_types_map[bench] = "No draw functions found"
                    
            except Exception as e:
                draw_types_map[bench] = f"Error parsing JSON: {str(e)[:50]}"
        else:
            draw_types_map[bench] = f"JSON file not found: {json_file}"
    
    return draw_types_map

def analyze_ftrace_files_graphite(folder_path, benches):
    """Analyze ftrace JSON files for Graphite backend (TODO feature)."""
    folder = Path(folder_path)
    draw_types_map = {}
    
    for bench in benches:
        json_file = folder / f"{bench}.json"
        
        if json_file.exists():
            # TODO: Implement Graphite-specific trace analysis
            draw_types_map[bench] = "TODO: Graphite trace analysis not yet implemented"
        else:
            draw_types_map[bench] = f"JSON file not found: {json_file}"
    
    return draw_types_map

def analyze_ftrace_files(folder_path, benches, folder_exists, folder_name):
    """Main dispatcher for ftrace analysis based on backend type."""
    if not folder_exists:
        print(f"\n⚠️  Skipping trace file analysis for '{folder_name}': Folder does not exist")
        return None
    
    backend_type = detect_backend_type(folder_path)
    print(f"  Backend type for '{folder_name}': {backend_type.upper()}")
    
    if backend_type == 'graphite':
        return analyze_ftrace_files_graphite(folder_path, benches)
    else:  # ganesh
        return analyze_ftrace_files_ganesh(folder_path, benches, folder_name)

def create_version_comparison_page(version_data, version_name, folder_paths, draw_types_maps, 
                                    summary_columns, missing_benchmarks_for_version, version_benchmarks):
    """Create comparison page for a specific version - only include benchmarks that exist in ALL backends."""
    dataframes = version_data['dataframes']
    
    # Find benchmarks that exist in ALL backends for this version
    backend_benchmarks = []
    for df in dataframes.values():
        backend_benchmarks.append(set(df['bench'].tolist()))
    
    if backend_benchmarks:
        # Find intersection of all benchmarks across all backends
        common_benchmarks = set.intersection(*backend_benchmarks)
        benches = sorted(list(common_benchmarks))
        print(f"    Filtered benchmarks: {len(common_benchmarks)} common out of {len(set.union(*backend_benchmarks))} total")
    else:
        benches = []
    
    if not benches:
        print(f"    WARNING: No common benchmarks found across all backends for version {version_name}")
        return pd.DataFrame()
    
    # Create ordered ID column
    ordered_ids = list(range(1, len(benches) + 1))
    
    # Prepare comparison data
    comparison_data = {
        'ID': ordered_ids,
        'Bench': benches
    }
    
    # Add mean columns for each backend (only for common benchmarks)
    backend_columns = {}  # Store column names for formula reference
    for col_name, df in sorted(dataframes.items()):
        mean_dict = dict(zip(df['bench'], df['mean']))
        
        # Remove everything starting from the last underscore
        display_name = col_name.rsplit('_', 1)[0] if '_' in col_name else col_name
        
        comparison_data[display_name] = [mean_dict.get(bench, float('nan')) for bench in benches]
        backend_columns[display_name] = display_name
    
    # Add ratio columns for this version (store as formula placeholders)
    ratio_configs = [
        ('grdawn_vk vs glesdmsaa', 'grdawn_vk', 'glesdmsaa'),
        ('vkdmsaa vs glesdmsaa', 'vkdmsaa', 'glesdmsaa'),
        ('grvk vs glesdmsaa', 'grvk', 'glesdmsaa')
    ]
    
    for ratio_name, num_backend, den_backend in ratio_configs:
        if num_backend in backend_columns and den_backend in backend_columns:
            # Store as formula placeholder
            comparison_data[ratio_name] = [f"FORMULA:{num_backend}/{den_backend}"] * len(benches)
    
    # Add summary columns from existing comparison page if provided
    if summary_columns:
        # Filter summary columns to only include common benchmarks
        for col_name, col_values in summary_columns.items():
            # Assuming summary columns are in the same order as original benches
            # For now, we'll just append as-is if length matches
            if len(col_values) == len(benches):
                comparison_data[col_name] = col_values
    
    # Add trace analysis summary columns from folders if provided (only for common benchmarks)
    if folder_paths and draw_types_maps:
        for idx, folder_path in enumerate(folder_paths):
            folder_name = Path(folder_path).name
            if idx < len(draw_types_maps) and draw_types_maps[idx] is not None:
                summary_col_name = f"trace_summary_of_{folder_name}"
                comparison_data[summary_col_name] = [draw_types_maps[idx].get(bench, "No trace data") for bench in benches]
    
    return pd.DataFrame(comparison_data)

def create_cross_version_page(version_groups, all_backends, summary_columns, missing_report, version_benchmarks):
    """Create cross-version comparison page - only include benchmarks that exist in ALL backends of ALL versions."""
    if len(version_groups) <= 1:
        return None
    
    # Find benchmarks that exist in ALL backends of ALL versions
    all_version_benchmarks = []
    
    for version_name, version_data in version_groups.items():
        # Get benchmarks for this version that exist in all backends of this version
        backend_benchmarks = []
        for df in version_data['dataframes'].values():
            backend_benchmarks.append(set(df['bench'].tolist()))
        
        if backend_benchmarks:
            version_common = set.intersection(*backend_benchmarks)
            all_version_benchmarks.append(version_common)
            print(f"    Version {version_name}: {len(version_common)} common benchmarks out of {len(set.union(*backend_benchmarks))} total")
    
    if all_version_benchmarks:
        # Find intersection across all versions
        common_benchmarks = set.intersection(*all_version_benchmarks)
        benches = sorted(list(common_benchmarks))
        print(f"    Cross-version common benchmarks: {len(common_benchmarks)} across all versions")
    else:
        benches = []
    
    if not benches:
        print(f"    WARNING: No common benchmarks found across all versions")
        return pd.DataFrame()
    
    # Prepare comparison data
    comparison_data = {
        'ID': list(range(1, len(benches) + 1)),
        'Bench': benches
    }
    
    # Add columns for each backend across versions (only for common benchmarks)
    versions = sorted(version_groups.keys())
    column_names = {}
    
    for backend in sorted(all_backends):
        for version in versions:
            col_name = f"{backend}_{version}"
            version_data = version_groups[version]
            found = False
            for data_col_name, df in version_data['dataframes'].items():
                if data_col_name.startswith(f"{backend}_") or data_col_name == backend:
                    mean_dict = dict(zip(df['bench'], df['mean']))
                    comparison_data[col_name] = [mean_dict.get(bench, float('nan')) for bench in benches]
                    column_names[col_name] = col_name
                    found = True
                    break
            if not found:
                comparison_data[col_name] = [float('nan')] * len(benches)
                column_names[col_name] = col_name
    
    # Add ratio columns comparing each version to the first version
    if len(versions) >= 2:
        base_version = versions[0]
        for backend in sorted(all_backends):
            base_col = f"{backend}_{base_version}"
            if base_col in column_names:
                for target_version in versions[1:]:
                    target_col = f"{backend}_{target_version}"
                    if target_col in column_names:
                        ratio_name = f"{backend}_{target_version}_vs_{base_version}"
                        comparison_data[ratio_name] = [f"FORMULA:{target_col}/{base_col}"] * len(benches)
    
    # Add summary columns from existing comparison page if provided
    if summary_columns:
        for col_name, col_values in summary_columns.items():
            if len(col_values) == len(benches):
                comparison_data[col_name] = col_values
    
    return pd.DataFrame(comparison_data)

def write_dataframe_with_formulas(writer, sheet_name, df):
    """Write dataframe to Excel with proper Excel formulas."""
    if df.empty:
        print(f"    WARNING: DataFrame for '{sheet_name}' is empty, skipping")
        return
    
    # First write the dataframe values without formulas
    df_for_write = df.copy()
    
    # Replace formula placeholders with None for initial write
    for col in df_for_write.columns:
        if df_for_write[col].dtype == 'object':
            if len(df_for_write) > 0:
                first_val = df_for_write[col].iloc[0]
                if isinstance(first_val, str) and first_val.startswith('FORMULA:'):
                    df_for_write[col] = None
    
    # Write the dataframe
    df_for_write.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Now add the formulas directly to the Excel sheet
    workbook = writer.book
    sheet = workbook[sheet_name]
    
    # Build column letter mapping
    col_letters = {}
    for idx, col_name in enumerate(df.columns, 1):
        col_letters[col_name] = get_column_letter(idx)
    
    # Add formulas to each column
    for col_idx, col_name in enumerate(df.columns, 1):
        if len(df) > 0:
            first_val = df[col_name].iloc[0]
            if isinstance(first_val, str) and first_val.startswith('FORMULA:'):
                formula_expr = first_val[8:]
                parts = formula_expr.split('/')
                if len(parts) == 2:
                    num_col_name = parts[0]
                    den_col_name = parts[1]
                    
                    if num_col_name in col_letters and den_col_name in col_letters:
                        num_col_letter = col_letters[num_col_name]
                        den_col_letter = col_letters[den_col_name]
                        
                        for row_idx in range(2, len(df) + 2):
                            formula = f"={num_col_letter}{row_idx}/{den_col_letter}{row_idx}"
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell.value = formula
                            cell.number_format = "0.000"
                            cell.comment = Comment(f"Formula: {num_col_name} / {den_col_name}", "Script")

def apply_table_formatting_to_sheet(sheet, df):
    """Apply Excel table formatting to a sheet."""
    if df.empty:
        return
    
    start_row = 1
    start_col = 1
    end_row = len(df) + 1
    end_col = len(df.columns)
    
    table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table_name = f"Table_{sheet.title.replace(' ', '_')[:20]}"
    
    for existing_table in list(sheet.tables.keys()):
        if existing_table.startswith("Table_"):
            del sheet.tables[existing_table]
    
    table = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.freeze_panes = sheet['B2']

def backup_original_sheets(writer, all_original_sheets):
    """Backup original backend sheets into the output workbook."""
    workbook = writer.book
    
    print(f"\n📋 Backing up original backend pages...")
    
    for sheet_name, df in all_original_sheets.items():
        clean_sheet_name = sheet_name[:31]
        final_sheet_name = clean_sheet_name
        counter = 1
        while final_sheet_name in workbook.sheetnames:
            final_sheet_name = f"{clean_sheet_name[:27]}_{counter}"
            counter += 1
        
        df.to_excel(writer, sheet_name=final_sheet_name, index=False)
        sheet = workbook[final_sheet_name]
        
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        sheet.freeze_panes = sheet['A2']
        print(f"  ✓ Backed up '{sheet_name}' -> sheet '{final_sheet_name}' ({len(df)} rows)")

def print_summary(folder_paths, folder_exists_list, draw_types_maps, version_groups, 
                  output_file, version_num, all_original_sheets, summary_columns, 
                  missing_report, unique_benchmarks, duplicate_report, 
                  missing_backends, single_version_backends):
    """Print a summary of the analysis."""
    print("\n" + "="*60)
    print("✅ Analysis complete!")
    
    if folder_paths:
        print("\n📁 Folders analyzed:")
        for idx, folder_path in enumerate(folder_paths):
            folder_name = Path(folder_path).name
            exists = folder_exists_list[idx] if idx < len(folder_exists_list) else False
            status = "✅ Analyzed" if exists else "⚠️  Skipped (not found)"
            print(f"  {idx + 1}. {folder_name}: {status}")
    else:
        print("\n📁 No folders provided - trace analysis skipped")
    
    print(f"\n📊 Versions processed: {len(version_groups)}")
    for version in version_groups.keys():
        num_backends = len(version_groups[version]['dataframes'])
        print(f"  - {version}: {num_backends} backends")
    
    # Print column/backend coverage summary
    if missing_backends:
        print(f"\n⚠️  Backends missing from some versions: {len(missing_backends)}")
        for backend, info in missing_backends.items():
            print(f"    - '{backend}' missing in: {', '.join(info['missing_in'])}")
    
    if single_version_backends:
        print(f"\n📌 Backends unique to single version: {len(single_version_backends)}")
    
    # Print duplicate benchmark summary
    if duplicate_report:
        total_duplicates = sum(len(sheets) for sheets in duplicate_report.values())
        print(f"\n⚠️  Duplicate Benchmarks Found: {total_duplicates} sheets with duplicates")
    
    # Print missing benchmark summary
    if missing_report:
        total_missing = sum(len(m) for m in missing_report.values())
        print(f"\n⚠️  Missing Benchmarks: {total_missing} total missing entries across versions")
    
    if unique_benchmarks:
        print(f"\n📌 Benchmarks unique to single version: {len(unique_benchmarks)}")
    
    print(f"\n📋 Original sheets backed up: {len(all_original_sheets)}")
    
    if summary_columns:
        print(f"\n📋 Summary columns extracted from existing comparison: {len(summary_columns)}")
    
    print(f"\n📁 Output file: {output_file}")
    print(f"   Version number: v{version_num}")
    
    print("\n📑 Sheets in output workbook:")
    print("  - [version]_comparison (one per API version, filtered to common benchmarks)")
    if len(version_groups) > 1:
        print("  - cross_version_comparison (filtered to benchmarks common across ALL versions)")
    print("  - backend_version (original data backups - unfiltered)")
    
    print("\n💡 Tips for using the Excel file:")
    print("  1. Each version page only includes benchmarks that exist in ALL backends of that version")
    print("  2. Cross-version page only includes benchmarks that exist in ALL backends of ALL versions")
    print("  3. Original backup sheets contain complete unfiltered data")
    print("  4. Use drop-down arrows in headers to sort/filter data")
    print("  5. First row and column are frozen for easy scrolling")

def main():
    """Main function to orchestrate the script."""
    print("="*60)
    print(f"Benchmark Analysis Tool - Version {VERSION}")
    print("="*60)
    
    # Validate arguments
    folder_paths, excel_files, folder_exists_list = validate_arguments()
    
    if folder_paths:
        print(f"\n📁 Folder paths ({len(folder_paths)}):")
        for folder_path in folder_paths:
            print(f"  - {folder_path}")
    else:
        print(f"\n📁 No folder paths provided - trace analysis will be skipped")
    
    print(f"\n📄 Excel files ({len(excel_files)}):")
    for excel_file in excel_files:
        print(f"  - {excel_file}")
    
    # Read all Excel files
    (all_dataframes, all_backends, version_info, version_groups, 
     all_original_sheets, summary_columns, missing_report, 
     unique_benchmarks, duplicate_report, missing_backends, 
     single_version_backends, version_benchmarks) = read_multiple_excel_files(excel_files)
    
    # Get unique benches for JSON analysis
    all_benches = set()
    for df in all_dataframes.values():
        all_benches.update(df['bench'].tolist())
    print(f"\n📋 Found {len(all_benches)} unique benchmarks")
    
    # Analyze ftrace files (if folders provided)
    draw_types_maps = []
    if folder_paths:
        print("\n🔍 Analyzing ftrace JSON files...")
        for idx, folder_path in enumerate(folder_paths):
            folder_name = Path(folder_path).name
            print(f"\n  Processing folder {idx + 1}: {folder_name}")
            folder_exists = folder_exists_list[idx] if idx < len(folder_exists_list) else False
            draw_types_map = analyze_ftrace_files(folder_path, all_benches, folder_exists, folder_name)
            draw_types_maps.append(draw_types_map)
    else:
        print("\n🔍 Skipping ftrace analysis - no folders provided")
    
    # Generate output filename with version number
    output_file, version_num = generate_output_filename(excel_files, folder_paths)
    
    # Write to Excel
    print(f"\n💾 Generating Excel workbook: {output_file} (version {version_num})")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Create comparison page for each version
        for version_name, version_data in version_groups.items():
            print(f"\n  Creating comparison page for version: {version_name}")
            missing_for_version = missing_report.get(version_name, set())
            version_df = create_version_comparison_page(
                version_data, version_name, folder_paths, draw_types_maps, 
                summary_columns, missing_for_version, version_benchmarks
            )
            
            if not version_df.empty:
                sheet_name = f"{version_name}_comparison"[:31]
                write_dataframe_with_formulas(writer, sheet_name, version_df)
                
                workbook = writer.book
                sheet = workbook[sheet_name]
                apply_table_formatting_to_sheet(sheet, version_df)
                print(f"    ✓ Created '{sheet_name}' with {len(version_df)} benchmarks, {len(version_df.columns)} columns")
            else:
                print(f"    ✗ Skipping '{version_name}_comparison' - no common benchmarks found")
        
        # Create cross-version comparison page if multiple versions
        if len(version_groups) > 1:
            print(f"\n  Creating cross-version comparison page")
            cross_version_df = create_cross_version_page(
                version_groups, all_backends, summary_columns, missing_report, version_benchmarks
            )
            if cross_version_df is not None and not cross_version_df.empty:
                write_dataframe_with_formulas(writer, 'cross_version_comparison', cross_version_df)
                
                workbook = writer.book
                sheet = workbook['cross_version_comparison']
                apply_table_formatting_to_sheet(sheet, cross_version_df)
                print(f"    ✓ Created 'cross_version_comparison' with {len(cross_version_df)} benchmarks, {len(cross_version_df.columns)} columns")
            else:
                print(f"    ✗ Skipping 'cross_version_comparison' - no common benchmarks across all versions")
        
        # Backup original backend sheets
        backup_original_sheets(writer, all_original_sheets)
    
    # Print summary
    print_summary(folder_paths, folder_exists_list, draw_types_maps, version_groups, 
                  output_file, version_num, all_original_sheets, summary_columns, 
                  missing_report, unique_benchmarks, duplicate_report, 
                  missing_backends, single_version_backends)
    
    print("\n" + "="*60)
    print(f"Benchmark Analysis Tool v{VERSION} - Execution Complete")
    print("="*60)

if __name__ == "__main__":
    main()